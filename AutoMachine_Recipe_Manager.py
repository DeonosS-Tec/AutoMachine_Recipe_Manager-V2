"""
Tea Recipe Management System v2
================================
Requirements: pip install openpyxl
Usage:        python tea_recipe_manager.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os, copy, json

# ─── Constants ────────────────────────────────────────────────────────────────

SIZES        = ["16oz", "22oz"]
OPTIONS      = ["Normal Ice", "Less Ice", "No Ice", "Hot"]
STD_SUGARS   = ["0%", "25%", "50%", "75%", "100%"]
INFO_HDRS    = ["Recipe Code", "Recipe name", "Product Code", "Recipe Group", "Recipe tags"]
FIXED_COLS   = ["Size", "Option", "Sugar Level"]
POPUP_HDR    = "Pop-up Tips"
OPTION_ORDER = {"Normal Ice": 0, "Less Ice": 1, "No Ice": 2, "Hot": 3}
INGREDIENTS_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ingredients.json")
CODE_MAPPINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "code_mappings.json")
GROUPS_TAGS_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "groups_tags.json")
UNDO_LIMIT   = 30

# Tags→Groups linked structure: {tag/country: [groups belonging to that country]}
DEFAULT_TAGS_GROUPS = {
    "Thai": [
        "TH-THAI TEA", "TH-SPARKLE", "TH-SMOOTHIE", "TH-MILK TEA",
        "TH-MATCHA", "TH-FRESH TEA", "TH-COLD", "TH-COCKTAIL",
        "TH-BUTTER MILK", "OH-SIGNATURE", "OH-MILK TEA", "OH-FRESH TEA",
        "OH-SIGNATURE TEST", "OH-MILK TEA TEST", "TEST"
    ],
    "Singapore": [
        "SG-THAI TEA", "SG-SPARKLE TEA", "SG-SMOOTHIE", "SG-MILK TEA",
        "SG-MATCHA", "SG-FRESH TEA", "SG-COLD", "SG-BUTTER MILK"
    ],
    "Australia": [],
    "Spain": [
        "THAI TEA-ESP", "SPARKLE TEA-ESP", "SLUSHIE-ESP", "MILK TEA-ESP",
        "MATCHA-ESP", "FRESH TEA-ESP", "FRUIT TEA-ESP", "COFFEE-ESP",
        "BUTTER MILK TEA-ESP"
    ],
    "USA": [
        "THAI TEA-USA", "SPARKLE TEA-USA", "SLUSHIE-USA", "MILK TEA-USA",
        "MATCHA-USA", "FRESH TEA-USA", "FRUIT TEA-USA", "COFFEE-USA",
        "BUTTER MILK TEA-USA"
    ]
}

# ─── QR Code Mappings (defaults — overridden by code_mappings.json) ───────────
# SI Code number determines sort order in QR export (SI0001 < SI0002 ... SI0006)
DEFAULT_SUGAR_MAP = {
    "0%":   "SI0001",
    "25%":  "SI0002",
    "50%":  "SI0003",
    "75%":  "SI0004",
    "100%": "SI0005",
    "10%":  "SI0006",
}
DEFAULT_ICE_MAP = {
    "No Ice":     "SI0028",
    "Less Ice":   "SI0029",
    "Normal Ice": "SI0030",
    "Hot":        "SI0031",
}
DEFAULT_SIZE_MAP = {
    "16oz": "SI0014",
    "22oz": "SI0015",
}

# Runtime maps — loaded from file or defaults
SUGAR_CODE_MAP = dict(DEFAULT_SUGAR_MAP)
ICE_CODE_MAP   = dict(DEFAULT_ICE_MAP)
SIZE_CODE_MAP  = dict(DEFAULT_SIZE_MAP)

QR_ORDER_PREFIX = "ORDERXXXX"
QR_SHEET_NAME   = "Finalized QR String"
QR_HEADERS      = [
    "Recipe Code", "Recipe name", "Size Code", "Size Name",
    "Ice Code", "Ice Level", "Sugar Code", "Sugar Level",
    "Finalized QR String"
]
QR_TABLE_STYLE  = "TableStyleLight21"

FONT_APTOS = Font(name="Aptos Narrow")
ALIGN_CTR  = Alignment(horizontal="center")
THIN       = Side(style="thin")
BORDER     = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# ─── Colour palette ───────────────────────────────────────────────────────────

BG         = "#FAF8F5"
SIDEBAR_BG = "#1C1A17"
PANEL_BG   = "#FFFFFF"
ACCENT     = "#C8973A"
TEXT_MAIN  = "#1C1A17"
TEXT_MUTED = "#7A6F5E"
TEXT_LIGHT = "#F0EBE3"
BORDER_CLR = "#E0D9CF"
BTN_ADD    = "#2E7D4F"
BTN_DEL    = "#B03A2E"
BTN_HOVER  = "#E8E0D4"
HEADER_BG  = "#2E2B25"
ROW_ALT    = "#F5F2EE"
MULTI_SEL_BG = "#5C4A2E"  # darker gold for secondary multi-selected rows in sidebar
ERR_BG     = "#FFE4E4"
ERR_FG     = "#B03A2E"

# ─── Helpers ──────────────────────────────────────────────────────────────────

def safe_name(name):
    for ch in [':', '\\', '/', '?', '*', '[', ']']:
        name = name.replace(ch, '')
    return (name or "Sheet")[:31]

def sugar_key(v):
    try:
        return float(str(v).replace('%', ''))
    except Exception:
        return 9999

def apply_cell(cell, value=None):
    if value is not None:
        cell.value = value
    cell.font      = FONT_APTOS
    cell.alignment = ALIGN_CTR
    cell.border    = BORDER

def col_width(ws, ci, pad=3):
    letter = get_column_letter(ci)
    mx = 0
    for row in ws.iter_rows(min_col=ci, max_col=ci):
        for c in row:
            if c.value is not None:
                mx = max(mx, len(str(c.value)))
    if mx:
        ws.column_dimensions[letter].width = mx + pad

def is_numeric(val):
    if val is None or str(val).strip() == "":
        return True
    try:
        float(str(val))
        return True
    except ValueError:
        return False

# ─── Ingredient Master List ───────────────────────────────────────────────────

DEFAULT_INGREDIENTS = {
    "J": [
        "J_Camellia Tea", "J_Guihua Tea", "J_White Peach Tea",
        "J_Jasmine Tea", "J_Water", "J_Milk", "J_Ice Hot",
        "J_Syrup mix Jiancha", "J_MilkMix", "J_Coconut water",
        "J_Pomegranate juice", "J_Whipping Cream", "J_Green grape juice",
        "J_Grape juice", "J_Black Tea", "J_Thai Tea", "J_Orange juice",
        "J_Unsweetened Condensed Milk", "J_Kalemix"
    ],
    "H": [
        "H_Lychee Tea", "H_Ta Hongpao Tea", "H_Maple Tea",
        "H_Sticky Rice Tea", "H_Rose Tea", "H_Water", "H_Milk",
        "H_Ice Hot", "H_Syrup mix Ohpolly", "H_Strawberry Tea",
        "H_Green Grape Tea", "H_Green grape juice", "H_Oat Milk"
    ]
}

def load_ingredients():
    if os.path.exists(INGREDIENTS_FILE):
        try:
            with open(INGREDIENTS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return copy.deepcopy(DEFAULT_INGREDIENTS)

def save_ingredients(data):
    try:
        with open(INGREDIENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("Error", f"Cannot save ingredients list:\n{e}")

def _tea_sort_key(name):
    """Sort key: Tea ingredients first, then others, alphabetically within each group."""
    return (0 if "tea" in name.lower() else 1, name.lower())

def flat_ingredients(data):
    """Return flat list: J_ group first, then H_, then others.
    Within each group: Tea ingredients first, then rest alphabetically."""
    def sort_group(items):
        return sorted(items, key=_tea_sort_key)

    j_items = sort_group(data.get("J", []))
    h_items = sort_group(data.get("H", []))
    others  = []
    for group in sorted(k for k in data if k not in ("J", "H")):
        others.extend(sort_group(data[group]))
    return j_items + h_items + others

def load_code_mappings():
    """Load Sugar/Ice/Size code mappings from JSON, fall back to defaults."""
    global SUGAR_CODE_MAP, ICE_CODE_MAP, SIZE_CODE_MAP
    if os.path.exists(CODE_MAPPINGS_FILE):
        try:
            with open(CODE_MAPPINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            SUGAR_CODE_MAP = data.get("sugar", dict(DEFAULT_SUGAR_MAP))
            ICE_CODE_MAP   = data.get("ice",   dict(DEFAULT_ICE_MAP))
            SIZE_CODE_MAP  = data.get("size",  dict(DEFAULT_SIZE_MAP))
            return
        except Exception:
            pass
    SUGAR_CODE_MAP = dict(DEFAULT_SUGAR_MAP)
    ICE_CODE_MAP   = dict(DEFAULT_ICE_MAP)
    SIZE_CODE_MAP  = dict(DEFAULT_SIZE_MAP)

def save_code_mappings():
    try:
        with open(CODE_MAPPINGS_FILE, "w", encoding="utf-8") as f:
            json.dump({"sugar": SUGAR_CODE_MAP,
                       "ice":   ICE_CODE_MAP,
                       "size":  SIZE_CODE_MAP},
                      f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("Error", f"Cannot save code mappings:\n{e}")

def si_code_sort_key(code_str):
    """Sort by SI code number: SI0001 < SI0002 ... regardless of label."""
    try:
        return int(code_str.replace("SI", "").lstrip("0") or "0")
    except Exception:
        return 9999

def load_groups_tags():
    """Returns dict {tag: [groups]}."""
    if os.path.exists(GROUPS_TAGS_FILE):
        try:
            with open(GROUPS_TAGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and "tags_groups" in data:
                return data["tags_groups"]
        except Exception:
            pass
    return copy.deepcopy(DEFAULT_TAGS_GROUPS)

def save_groups_tags(tags_groups):
    """Save dict {tag: [groups]}."""
    try:
        with open(GROUPS_TAGS_FILE, "w", encoding="utf-8") as f:
            json.dump({"tags_groups": tags_groups}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("Error", f"Cannot save groups/tags:\n{e}")

# ─── Data model ───────────────────────────────────────────────────────────────

SYRUP_LAST_NAME = "J_Syrup mix Jiancha"   # exact name to always place last

def _syrup_last(ingredients):
    """Return ingredients list with SYRUP_LAST_NAME moved to last position."""
    if SYRUP_LAST_NAME not in ingredients:
        return ingredients
    others = [i for i in ingredients if i != SYRUP_LAST_NAME]
    return others + [SYRUP_LAST_NAME]

def _reorder_rows_to_ingredients(rd):
    """
    Re-number Ingredient/Usage keys in all rows to match rd.ingredients order.
    Call after changing ingredient order so row data stays consistent.
    """
    old_order = []
    # Detect current order from first non-empty row
    if rd.rows:
        first = rd.rows[0]
        i = 1
        while f"Ingredient{i}" in first:
            old_order.append(str(first.get(f"Ingredient{i}", "") or ""))
            i += 1

    if not old_order or old_order == rd.ingredients:
        return  # No reorder needed

    for row in rd.rows:
        # Read current values
        old_vals = {}
        for oi, name in enumerate(old_order, 1):
            old_vals[name] = {
                "ingr":  row.pop(f"Ingredient{oi}", ""),
                "usage": row.pop(f"Usage{oi}", ""),
            }
        # Write in new order
        for ni, name in enumerate(rd.ingredients, 1):
            v = old_vals.get(name, {"ingr": name, "usage": ""})
            row[f"Ingredient{ni}"] = v["ingr"]
            row[f"Usage{ni}"]      = v["usage"]

class RecipeData:
    def __init__(self, name="New Recipe"):
        self.name         = name
        self.recipe_code  = ""
        self.product_code = ""
        self.group        = ""
        self.tags         = ""
        self.ingredients  = []   # list of str: "Ingredient1", "Ingredient2", ...
        self.rows         = []   # list of dicts

    @classmethod
    def from_sheet(cls, ws):
        rd = cls(ws.title)
        rd.recipe_code  = ws.cell(2, 1).value or ""
        rd.name         = ws.cell(2, 2).value or ws.title
        rd.product_code = ws.cell(2, 3).value or ""
        rd.group        = ws.cell(2, 4).value or ""
        rd.tags         = ws.cell(2, 5).value or ""

        headers = {}
        for cell in ws[3]:
            if cell.value:
                headers[cell.column] = cell.value

        # Collect Ingredient column indices in order
        ingr_col_indices = sorted(
            [(ci, name) for ci, name in headers.items()
             if str(name).startswith("Ingredient")],
            key=lambda x: x[0]
        )
        n_ingr = len(ingr_col_indices)

        for src_row in ws.iter_rows(min_row=4):
            row_dict = {}
            for cell in src_row:
                hdr = headers.get(cell.column)
                if hdr:
                    row_dict[hdr] = cell.value
            if "Ice" in row_dict:
                row_dict["Option"] = row_dict.pop("Ice")
            if any(v not in (None, "") for v in row_dict.values()):
                rd.rows.append(row_dict)

        # Extract actual ingredient name values from first data row
        # e.g. row["Ingredient1"] = "J_Black Tea"  (the cell value, not the column header)
        if rd.rows:
            first = rd.rows[0]
            rd.ingredients = [
                str(first.get(f"Ingredient{i}", "") or "")
                for i in range(1, n_ingr + 1)
            ]
        else:
            rd.ingredients = [""] * n_ingr

        # Fix 3: move "J_Syrup mix Jiancha" to last position
        rd.ingredients = _syrup_last(rd.ingredients)
        # Re-number ingredient keys in all rows to match new order
        if len(rd.rows) > 0:
            _reorder_rows_to_ingredients(rd)

        return rd

    def to_sheet_data(self):
        # Fix 3: ensure syrup is last before export
        self.ingredients = _syrup_last(self.ingredients)
        _reorder_rows_to_ingredients(self)

        n = len(self.ingredients)
        ingr_cols = []
        for i in range(1, n + 1):
            ingr_cols.append(f"Ingredient{i}")
            ingr_cols.append(f"Usage{i}")
        out_cols = FIXED_COLS + ingr_cols + [POPUP_HDR]
        sorted_rows = sorted(
            self.rows,
            key=lambda d: (
                str(d.get("Size", "")),
                OPTION_ORDER.get(str(d.get("Option", "")), 99),
                sugar_key(d.get("Sugar Level", ""))
            )
        )
        return out_cols, sorted_rows

    def deep_copy(self):
        rd = RecipeData(self.name)
        rd.recipe_code  = self.recipe_code
        rd.product_code = self.product_code
        rd.group        = self.group
        rd.tags         = self.tags
        rd.ingredients  = self.ingredients[:]
        rd.rows         = copy.deepcopy(self.rows)
        return rd

# ─── Excel I/O ───────────────────────────────────────────────────────────────

def load_workbook_data(path):
    wb = load_workbook(path)
    recipes = []
    for name in wb.sheetnames:
        ws = wb[name]
        recipes.append(RecipeData.from_sheet(ws))
    wb.close()
    return recipes

def save_workbook(recipes, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for rd in recipes:
        title = safe_name(rd.name)
        ws = wb.create_sheet(title=title)
        for ci, hdr in enumerate(INFO_HDRS, 1):
            apply_cell(ws.cell(1, ci), hdr)
        vals = [rd.recipe_code, rd.name, rd.product_code, rd.group, rd.tags]
        for ci, v in enumerate(vals, 1):
            apply_cell(ws.cell(2, ci), v)
        out_cols, sorted_rows = rd.to_sheet_data()
        for ci, hdr in enumerate(out_cols, 1):
            apply_cell(ws.cell(3, ci), hdr)
        for ri, row in enumerate(sorted_rows, 4):
            for ci, hdr in enumerate(out_cols, 1):
                apply_cell(ws.cell(ri, ci), row.get(hdr, ""))
        for ci in range(1, len(out_cols) + 1):
            col_width(ws, ci)
    wb.save(path)

def export_qr_workbook(recipes, out_path, ingr_master):
    """
    Export all recipes to a QR Code Excel file matching the template structure.
    - Recipes sorted by recipe_code ascending
    - Sugar sorted by SI code number (SI0001→SI0006 in order)
    - Table styled as TableStyleLight21 (matches template)
    - Column widths auto-fit to content
    Returns list of warning strings for unmapped codes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = QR_SHEET_NAME

    data_font = Font(name="Aptos Narrow", size=11)
    warnings  = []

    # Write header row
    for ci, hdr in enumerate(QR_HEADERS, 1):
        cell      = ws.cell(1, ci, hdr)
        cell.font = data_font

    # Sort recipes by recipe_code ascending
    sorted_recipes = sorted(recipes, key=lambda rd: str(rd.recipe_code or ""))

    row_num = 2
    for rd in sorted_recipes:
        # Sort rows: Size → Option → Sugar by SI code number
        def row_sort_key(d):
            size   = str(d.get("Size", ""))
            option = str(d.get("Option", ""))
            sugar  = str(d.get("Sugar Level", ""))
            size_si   = si_code_sort_key(SIZE_CODE_MAP.get(size, "SI9999"))
            option_si = si_code_sort_key(ICE_CODE_MAP.get(option, "SI9999"))
            sugar_si  = si_code_sort_key(SUGAR_CODE_MAP.get(sugar, "SI9999"))
            return (size_si, option_si, sugar_si)

        for row in sorted(rd.rows, key=row_sort_key):
            size   = str(row.get("Size", ""))
            option = str(row.get("Option", ""))
            sugar  = str(row.get("Sugar Level", ""))

            size_code  = SIZE_CODE_MAP.get(size, "")
            ice_code   = ICE_CODE_MAP.get(option, "")
            sugar_code = SUGAR_CODE_MAP.get(sugar, "")

            if size   and not size_code:
                warnings.append(f'[{rd.name}] Unknown size: "{size}" — no Size Code')
            if option and not ice_code:
                warnings.append(f'[{rd.name}] Unknown ice: "{option}" — no Ice Code')
            if sugar  and not sugar_code:
                warnings.append(f'[{rd.name}] Unknown sugar: "{sugar}" — no Sugar Code')

            qr_string = f"{QR_ORDER_PREFIX}|{rd.recipe_code}|{size_code},{ice_code},{sugar_code}|"

            for ci, val in enumerate([
                rd.recipe_code, rd.name,
                size_code, size,
                ice_code, option,
                sugar_code, sugar,
                qr_string,
            ], 1):
                cell      = ws.cell(row_num, ci, val)
                cell.font = data_font

            row_num += 1

    last_data_row = row_num - 1

    # Auto-fit column widths to longest content (header or data)
    for ci, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=last_data_row), 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if max_len:
            ws.column_dimensions[get_column_letter(ci)].width = max_len + 3

    # Apply Format as Table — TableStyleLight21 (matches template exactly)
    if last_data_row >= 1:
        tbl_ref = f"A1:{get_column_letter(len(QR_HEADERS))}{last_data_row}"
        tbl = Table(displayName="QRTable", ref=tbl_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name=QR_TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)

    wb.save(out_path)
    return warnings

# ─── Search Combobox ──────────────────────────────────────────────────────────

class SearchCombobox(tk.Frame):
    VISIBLE_ROWS = 10

    def __init__(self, parent, items, width=18, on_select=None, bg=None, **kw):
        bg = bg or PANEL_BG
        super().__init__(parent, bg=bg)
        self._all_items  = items
        self._on_select  = on_select
        self._listbox    = None
        self._toplevel   = None
        self._var        = tk.StringVar()
        self._width      = width
        self._suppress   = False
        self._entry = tk.Entry(self, textvariable=self._var, width=width,
                               font=("TkDefaultFont", 9),
                               bg=BG, fg=TEXT_MAIN,
                               insertbackground=TEXT_MAIN,
                               bd=1, relief="solid")
        self._entry.pack(fill="x")

        self._var.trace_add("write", self._on_type)
        # Open dropdown ONLY on explicit mouse click — never on FocusIn (avoids auto-open on tab switch/render)
        self._entry.bind("<Button-1>", self._show_all)
        self._entry.bind("<FocusOut>", self._on_focus_out)
        self._entry.bind("<Return>",   self._on_return)
        self._entry.bind("<Escape>",   self._close_dropdown)
        self._entry.bind("<Down>",     self._focus_list)
        self._entry.bind("<MouseWheel>", self._scroll_list)

    def get(self):
        return self._var.get()

    def set(self, val):
        self._suppress = True
        self._var.set(val)
        self._suppress = False

    def update_items(self, items):
        self._all_items = items

    def _show_all(self, event=None):
        self._open_dropdown(self._all_items)

    def _on_type(self, *_):
        if self._suppress:
            return
        q = self._var.get().lower()
        filtered = [i for i in self._all_items if q in i.lower()] if q else self._all_items
        self._open_dropdown(filtered)

    def _open_dropdown(self, items):
        self._close_dropdown()
        if not items:
            return
        self._toplevel = tk.Toplevel(self)
        self._toplevel.wm_overrideredirect(True)
        self._toplevel.attributes("-topmost", True)

        x = self._entry.winfo_rootx()
        y = self._entry.winfo_rooty() + self._entry.winfo_height()
        w = max(self._entry.winfo_width(), 200)
        h = min(len(items), self.VISIBLE_ROWS) * 22 + 4
        self._toplevel.geometry(f"{w}x{h}+{x}+{y}")

        frame = tk.Frame(self._toplevel, bg=PANEL_BG, bd=1, relief="solid")
        frame.pack(fill="both", expand=True)

        sb = tk.Scrollbar(frame)
        sb.pack(side="right", fill="y")

        self._listbox = tk.Listbox(frame, yscrollcommand=sb.set,
                                   font=("TkDefaultFont", 9),
                                   bg=PANEL_BG, fg=TEXT_MAIN,
                                   selectbackground=ACCENT,
                                   selectforeground="#FFF",
                                   activestyle="none",
                                   bd=0, highlightthickness=0,
                                   height=min(len(items), self.VISIBLE_ROWS))
        self._listbox.pack(fill="both", expand=True)
        sb.config(command=self._listbox.yview)

        for item in items:
            self._listbox.insert("end", item)

        self._listbox.bind("<ButtonRelease-1>", self._pick_selected)
        self._listbox.bind("<Return>",          self._pick_selected)
        self._listbox.bind("<Escape>",          self._close_dropdown)
        # Fix 2: mouse wheel works on listbox
        self._listbox.bind("<MouseWheel>", lambda e:
            self._listbox.yview_scroll(int(-1 * e.delta / 120), "units"))

    def _pick_selected(self, event=None):
        if not self._listbox:
            return
        sel = self._listbox.curselection()
        if sel:
            val = self._listbox.get(sel[0])
            self._suppress = True
            self._var.set(val)
            self._suppress = False
            if self._on_select:
                self._on_select(val)
        self._close_dropdown()

    # Fix 2: Enter keeps current typed value, don't jump to first item
    def _on_return(self, event=None):
        if self._listbox and self._listbox.size() > 0:
            sel = self._listbox.curselection()
            if sel:
                # User navigated list with arrow → pick highlighted
                self._pick_selected()
            else:
                # User just typed and pressed Enter → keep what they typed, close
                if self._on_select:
                    self._on_select(self._var.get())
                self._close_dropdown()
        else:
            self._close_dropdown()

    def _focus_list(self, event=None):
        if self._listbox:
            self._listbox.focus_set()
            self._listbox.selection_set(0)

    def _scroll_list(self, event=None):
        if self._listbox:
            self._listbox.yview_scroll(int(-1 * event.delta / 120), "units")

    def _on_focus_out(self, event=None):
        # Fix 10: close on focus-out with slight delay (allows listbox click to register)
        self.after(200, self._close_if_not_focused)

    def _close_if_not_focused(self):
        """Close dropdown if focus has moved away from both entry and listbox."""
        try:
            focused = self.focus_get()
        except Exception:
            focused = None
        # Check if focused widget belongs to this combobox or its dropdown
        if focused is self._entry:
            return
        if self._listbox and focused is self._listbox:
            return
        self._close_dropdown()

    def _close_dropdown(self, event=None):
        if self._toplevel:
            try:
                self._toplevel.destroy()
            except Exception:
                pass
            self._toplevel = None
            self._listbox  = None

# ─── Simple Input Dialog ──────────────────────────────────────────────────────

class SimpleInput(tk.Toplevel):
    def __init__(self, parent, title, prompt):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.grab_set()
        self.resizable(False, False)
        tk.Label(self, text=prompt, pady=8, padx=12).pack()
        self._var = tk.StringVar()
        e = tk.Entry(self, textvariable=self._var, width=24)
        e.pack(padx=12, pady=4)
        e.focus_set()
        f = tk.Frame(self)
        f.pack(pady=8)
        tk.Button(f, text="OK",     command=self._ok,     width=8).pack(side="left", padx=4)
        tk.Button(f, text="Cancel", command=self.destroy, width=8).pack(side="left")
        self.bind("<Return>", lambda e: self._ok())
        self.wait_window()

    def _ok(self):
        self.result = self._var.get()
        self.destroy()

# ─── Groups & Tags Manager Dialog ────────────────────────────────────────────

class GroupsTagsDialog(tk.Toplevel):
    """Manage linked Tags→Groups structure: {tag/country: [groups]}."""

    def __init__(self, parent, tags_groups, on_save):
        super().__init__(parent)
        self.title("Manage Groups & Tags")
        self.geometry("560x460")
        self.configure(bg=BG)
        self.grab_set()
        self._data    = copy.deepcopy(tags_groups)
        self._on_save = on_save
        self._tag_var = tk.StringVar()
        self._build()
        self._refresh_all()

    def _build(self):
        tk.Label(self, text="Groups & Tags Manager",
                 font=("Georgia", 13, "bold"),
                 bg=HEADER_BG, fg=TEXT_LIGHT, pady=12).pack(fill="x")

        # Left: Tags list | Right: Groups for selected tag
        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True, padx=16, pady=8)
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=2)

        # ── Left: Tags ────────────────────────────────────────────────
        left = tk.Frame(main, bg=BG)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        tk.Label(left, text="Tags (Countries)", font=("TkDefaultFont", 9, "bold"),
                 bg=BG, fg=TEXT_MAIN).pack(anchor="w")

        lf_t = tk.Frame(left, bg=BG)
        lf_t.pack(fill="both", expand=True)
        sb_t = tk.Scrollbar(lf_t)
        sb_t.pack(side="right", fill="y")
        self._tag_lb = tk.Listbox(lf_t, yscrollcommand=sb_t.set,
                                   font=("TkDefaultFont", 10),
                                   bg=PANEL_BG, fg=TEXT_MAIN,
                                   selectbackground=ACCENT, selectforeground="#FFF",
                                   bd=1, relief="solid", activestyle="none", height=10)
        self._tag_lb.pack(fill="both", expand=True)
        sb_t.config(command=self._tag_lb.yview)
        self._tag_lb.bind("<<ListboxSelect>>", self._on_tag_select)

        add_tag_f = tk.Frame(left, bg=BG, pady=4)
        add_tag_f.pack(fill="x")
        self._new_tag_var = tk.StringVar()
        tk.Entry(add_tag_f, textvariable=self._new_tag_var, width=14,
                 font=("TkDefaultFont", 9), bd=1, relief="solid", bg=PANEL_BG).pack(side="left")
        tk.Button(add_tag_f, text="＋", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=6, pady=3,
                  cursor="hand2", command=self._add_tag).pack(side="left", padx=4)
        tk.Button(add_tag_f, text="✕", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_DEL, bd=0, padx=6, pady=3,
                  cursor="hand2", command=self._del_tag).pack(side="left")

        # ── Right: Groups for selected tag ────────────────────────────
        right = tk.Frame(main, bg=BG)
        right.grid(row=0, column=1, sticky="nsew")

        self._grp_title = tk.Label(right, text="Groups (select a tag first)",
                                    font=("TkDefaultFont", 9, "bold"),
                                    bg=BG, fg=TEXT_MAIN)
        self._grp_title.pack(anchor="w")

        lf_g = tk.Frame(right, bg=BG)
        lf_g.pack(fill="both", expand=True)
        sb_g = tk.Scrollbar(lf_g)
        sb_g.pack(side="right", fill="y")
        self._grp_lb = tk.Listbox(lf_g, yscrollcommand=sb_g.set,
                                   font=("TkDefaultFont", 10),
                                   bg=PANEL_BG, fg=TEXT_MAIN,
                                   selectbackground=ACCENT, selectforeground="#FFF",
                                   bd=1, relief="solid", activestyle="none", height=10)
        self._grp_lb.pack(fill="both", expand=True)
        sb_g.config(command=self._grp_lb.yview)

        add_grp_f = tk.Frame(right, bg=BG, pady=4)
        add_grp_f.pack(fill="x")
        self._new_grp_var = tk.StringVar()
        tk.Entry(add_grp_f, textvariable=self._new_grp_var, width=20,
                 font=("TkDefaultFont", 9), bd=1, relief="solid", bg=PANEL_BG).pack(side="left")
        tk.Button(add_grp_f, text="＋ Add Group", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=6, pady=3,
                  cursor="hand2", command=self._add_grp).pack(side="left", padx=4)
        tk.Button(add_grp_f, text="✕", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_DEL, bd=0, padx=6, pady=3,
                  cursor="hand2", command=self._del_grp).pack(side="left")

        foot = tk.Frame(self, bg=BG, pady=10, padx=16)
        foot.pack(fill="x")
        tk.Button(foot, text="💾  Save & Close",
                  font=("TkDefaultFont", 10, "bold"),
                  fg="#1C1A17", bg=ACCENT, bd=0, padx=16, pady=6,
                  cursor="hand2", command=self._save).pack(side="right")
        tk.Button(foot, text="Cancel",
                  font=("TkDefaultFont", 10),
                  fg=TEXT_MAIN, bg=BG, bd=1, relief="solid",
                  padx=12, pady=6, cursor="hand2",
                  command=self.destroy).pack(side="right", padx=8)

    def _refresh_all(self):
        self._tag_lb.delete(0, "end")
        for tag in self._data:
            self._tag_lb.insert("end", tag)
        self._grp_lb.delete(0, "end")

    def _on_tag_select(self, event=None):
        sel = self._tag_lb.curselection()
        if not sel:
            return
        tag = self._tag_lb.get(sel[0])
        self._tag_var.set(tag)
        self._grp_title.config(text=f"Groups for: {tag}")
        self._grp_lb.delete(0, "end")
        for grp in self._data.get(tag, []):
            self._grp_lb.insert("end", grp)

    def _current_tag(self):
        return self._tag_var.get()

    def _add_tag(self):
        val = self._new_tag_var.get().strip()
        if val and val not in self._data:
            self._data[val] = []
            self._refresh_all()
        self._new_tag_var.set("")

    def _del_tag(self):
        sel = self._tag_lb.curselection()
        if sel:
            tag = self._tag_lb.get(sel[0])
            if messagebox.askyesno("Delete Tag", f"Delete tag '{tag}' and all its groups?",
                                   parent=self):
                self._data.pop(tag, None)
                self._tag_var.set("")
                self._grp_title.config(text="Groups (select a tag first)")
                self._refresh_all()

    def _add_grp(self):
        tag = self._current_tag()
        val = self._new_grp_var.get().strip()
        if not tag:
            messagebox.showwarning("Select Tag", "Please select a tag first.", parent=self)
            return
        if val and val not in self._data.get(tag, []):
            self._data.setdefault(tag, []).append(val)
            self._on_tag_select()
        self._new_grp_var.set("")

    def _del_grp(self):
        tag = self._current_tag()
        sel = self._grp_lb.curselection()
        if tag and sel:
            grp = self._grp_lb.get(sel[0])
            if grp in self._data.get(tag, []):
                self._data[tag].remove(grp)
            self._on_tag_select()

    def _save(self):
        self._on_save(self._data)
        self.destroy()

# ─── Code Mapping Manager Dialog ─────────────────────────────────────────────

class CodeMappingDialog(tk.Toplevel):
    """
    Manage Sugar / Ice / Size code mappings.
    Each tab: list of (label → SI Code) pairs. Add / delete with in-use warning.
    Saved to code_mappings.json immediately on Save.
    """
    TAB_DEFS = [
        ("Sugar",    "sugar",  SUGAR_CODE_MAP, "e.g. 30%",  "e.g. SI0007"),
        ("Ice",      "ice",    ICE_CODE_MAP,   "e.g. Extra Ice", "e.g. SI0032"),
        ("Size",     "size",   SIZE_CODE_MAP,  "e.g. 32oz", "e.g. SI0016"),
    ]

    def __init__(self, parent, recipes, on_save):
        super().__init__(parent)
        self.title("Code Mapping Manager")
        self.geometry("600x480")
        self.configure(bg=BG)
        self.grab_set()
        self._recipes  = recipes
        self._on_save  = on_save
        # Deep copy current maps to edit locally
        self._maps = {
            "sugar": dict(SUGAR_CODE_MAP),
            "ice":   dict(ICE_CODE_MAP),
            "size":  dict(SIZE_CODE_MAP),
        }
        self._build()

    def _build(self):
        tk.Label(self, text="Code Mapping Manager",
                 font=("Georgia", 13, "bold"),
                 bg=HEADER_BG, fg=TEXT_LIGHT, pady=12).pack(fill="x")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=16, pady=8)

        self._tab_widgets = {}
        for tab_label, key, _, lbl_hint, code_hint in self.TAB_DEFS:
            f = tk.Frame(nb, bg=BG)
            nb.add(f, text=f"  {tab_label}  ")
            self._build_tab(f, key, lbl_hint, code_hint)
            self._tab_widgets[key] = f

        foot = tk.Frame(self, bg=BG, pady=10, padx=16)
        foot.pack(fill="x")
        tk.Button(foot, text="💾  Save & Close",
                  font=("TkDefaultFont", 10, "bold"),
                  fg="#1C1A17", bg=ACCENT, bd=0, padx=16, pady=6,
                  cursor="hand2", command=self._save).pack(side="right")
        tk.Button(foot, text="Cancel",
                  font=("TkDefaultFont", 10),
                  fg=TEXT_MAIN, bg=BG, bd=1, relief="solid",
                  padx=12, pady=6, cursor="hand2",
                  command=self.destroy).pack(side="right", padx=8)

    def _build_tab(self, parent, key, lbl_hint, code_hint):
        # List frame
        lf = tk.Frame(parent, bg=BG)
        lf.pack(fill="both", expand=True, padx=8, pady=(8, 4))

        cols = ("Label", "SI Code")
        tv = ttk.Treeview(lf, columns=cols, show="headings", height=10)
        tv.heading("Label",   text="Label")
        tv.heading("SI Code", text="SI Code")
        tv.column("Label",   width=220, anchor="w")
        tv.column("SI Code", width=160, anchor="center")

        sb = tk.Scrollbar(lf, command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        tv.pack(fill="both", expand=True)

        self._refresh_tv(tv, key)

        # Add row
        add_f = tk.Frame(parent, bg=BG, padx=8, pady=4)
        add_f.pack(fill="x")

        lbl_var  = tk.StringVar()
        code_var = tk.StringVar()

        tk.Label(add_f, text="Label:", bg=BG, fg=TEXT_MAIN,
                 font=("TkDefaultFont", 9)).pack(side="left")
        tk.Entry(add_f, textvariable=lbl_var, width=14,
                 font=("TkDefaultFont", 9), bd=1, relief="solid", bg=PANEL_BG).pack(
            side="left", padx=(4, 8))

        tk.Label(add_f, text="SI Code:", bg=BG, fg=TEXT_MAIN,
                 font=("TkDefaultFont", 9)).pack(side="left")
        tk.Entry(add_f, textvariable=code_var, width=10,
                 font=("TkDefaultFont", 9), bd=1, relief="solid", bg=PANEL_BG).pack(
            side="left", padx=(4, 8))

        def add_item():
            lbl  = lbl_var.get().strip()
            code = code_var.get().strip()
            if not lbl or not code:
                messagebox.showwarning("Missing", "Please enter both Label and SI Code.",
                                       parent=self)
                return
            if lbl in self._maps[key]:
                messagebox.showwarning("Duplicate",
                                       f'"{lbl}" already exists in this mapping.',
                                       parent=self)
                return
            self._maps[key][lbl] = code
            self._refresh_tv(tv, key)
            lbl_var.set(""); code_var.set("")

        def del_item():
            sel = tv.selection()
            if not sel:
                return
            lbl = tv.item(sel[0], "values")[0]
            # Check if label is used in any recipe row
            used_in = []
            for rd in self._recipes:
                if key == "sugar":
                    if any(str(r.get("Sugar Level","")) == lbl for r in rd.rows):
                        used_in.append(rd.name)
                elif key == "ice":
                    if any(str(r.get("Option","")) == lbl for r in rd.rows):
                        used_in.append(rd.name)
                elif key == "size":
                    if any(str(r.get("Size","")) == lbl for r in rd.rows):
                        used_in.append(rd.name)
            if used_in:
                ans = messagebox.askyesno(
                    "In Use — Delete Anyway?",
                    f'"{lbl}" is used in recipe(s):\n  ' +
                    "\n  ".join(used_in[:10]) +
                    "\n\nDeleting this mapping will leave those rows without a QR code.\n"
                    "Delete anyway?",
                    parent=self
                )
                if not ans:
                    return
            self._maps[key].pop(lbl, None)
            self._refresh_tv(tv, key)

        tk.Button(add_f, text="＋ Add", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=8, pady=4,
                  cursor="hand2", command=add_item).pack(side="left")
        tk.Button(add_f, text="✕ Delete Selected", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_DEL, bd=0, padx=8, pady=4,
                  cursor="hand2", command=del_item).pack(side="left", padx=6)

    def _refresh_tv(self, tv, key):
        tv.delete(*tv.get_children())
        for lbl, code in sorted(self._maps[key].items(),
                                 key=lambda x: si_code_sort_key(x[1])):
            tv.insert("", "end", values=(lbl, code))

    def _save(self):
        global SUGAR_CODE_MAP, ICE_CODE_MAP, SIZE_CODE_MAP
        SUGAR_CODE_MAP = dict(self._maps["sugar"])
        ICE_CODE_MAP   = dict(self._maps["ice"])
        SIZE_CODE_MAP  = dict(self._maps["size"])
        save_code_mappings()
        self._on_save()
        self.destroy()

# ─── Ingredient Manager Dialog ────────────────────────────────────────────────

class IngredientManagerDialog(tk.Toplevel):
    def __init__(self, parent, ingr_data, on_save):
        super().__init__(parent)
        self.title("Ingredient Master List")
        self.geometry("560x520")
        self.configure(bg=BG)
        self.grab_set()
        self._data    = copy.deepcopy(ingr_data)
        self._on_save = on_save
        self._build()
        self._refresh()

    def _build(self):
        tk.Label(self, text="Ingredient Master List",
                 font=("Georgia", 13, "bold"),
                 bg=HEADER_BG, fg=TEXT_LIGHT, pady=12).pack(fill="x")

        top = tk.Frame(self, bg=BG, pady=8, padx=16)
        top.pack(fill="x")

        tk.Label(top, text="Group:", font=("TkDefaultFont", 10, "bold"),
                 bg=BG, fg=TEXT_MAIN).pack(side="left")

        self._group_var = tk.StringVar()
        self._group_cb  = ttk.Combobox(top, textvariable=self._group_var, width=10, state="readonly")
        self._group_cb.pack(side="left", padx=8)
        self._group_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_list())

        tk.Button(top, text="＋ New Group", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=8, pady=4,
                  cursor="hand2", command=self._add_group).pack(side="left", padx=4)
        tk.Button(top, text="✕ Del Group", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_DEL, bd=0, padx=8, pady=4,
                  cursor="hand2", command=self._del_group).pack(side="left")

        lf = tk.Frame(self, bg=BG, padx=16)
        lf.pack(fill="both", expand=True)

        sb = tk.Scrollbar(lf)
        sb.pack(side="right", fill="y")
        self._lb = tk.Listbox(lf, yscrollcommand=sb.set,
                              font=("TkDefaultFont", 10),
                              bg=PANEL_BG, fg=TEXT_MAIN,
                              selectbackground=ACCENT, selectforeground="#FFF",
                              bd=1, relief="solid", activestyle="none")
        self._lb.pack(fill="both", expand=True)
        sb.config(command=self._lb.yview)

        add_f = tk.Frame(self, bg=BG, padx=16, pady=8)
        add_f.pack(fill="x")
        self._new_var = tk.StringVar()
        tk.Entry(add_f, textvariable=self._new_var, width=30,
                 font=("TkDefaultFont", 10), bd=1, relief="solid", bg=BG).pack(side="left")
        tk.Button(add_f, text="＋ Add", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=10, pady=4,
                  cursor="hand2", command=self._add_item).pack(side="left", padx=6)
        tk.Button(add_f, text="✕ Remove Selected", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_DEL, bd=0, padx=10, pady=4,
                  cursor="hand2", command=self._del_item).pack(side="left")

        foot = tk.Frame(self, bg=BG, pady=10, padx=16)
        foot.pack(fill="x")
        tk.Button(foot, text="💾  Save & Close",
                  font=("TkDefaultFont", 10, "bold"),
                  fg="#1C1A17", bg=ACCENT, bd=0, padx=16, pady=6,
                  cursor="hand2", command=self._save).pack(side="right")
        tk.Button(foot, text="Cancel",
                  font=("TkDefaultFont", 10),
                  fg=TEXT_MAIN, bg=BG, bd=1, relief="solid",
                  padx=12, pady=6, cursor="hand2",
                  command=self.destroy).pack(side="right", padx=8)

    def _refresh(self):
        groups = sorted(self._data.keys())
        self._group_cb["values"] = groups
        if groups and self._group_var.get() not in groups:
            self._group_var.set(groups[0])
        self._refresh_list()

    def _refresh_list(self):
        self._lb.delete(0, "end")
        g = self._group_var.get()
        if g in self._data:
            for item in sorted(self._data[g]):
                self._lb.insert("end", item)

    def _add_group(self):
        dlg = SimpleInput(self, "New Group", "Enter group prefix (e.g. K):")
        name = dlg.result
        if name:
            name = name.strip()
            if name and name not in self._data:
                self._data[name] = []
                self._group_var.set(name)
                self._refresh()

    def _del_group(self):
        g = self._group_var.get()
        if not g:
            return
        if not messagebox.askyesno("Delete Group",
                                   f"Delete group '{g}' and all its ingredients?",
                                   parent=self):
            return
        self._data.pop(g, None)
        self._refresh()

    def _add_item(self):
        g   = self._group_var.get()
        val = self._new_var.get().strip()
        if not g or not val:
            return
        if val not in self._data[g]:
            self._data[g].append(val)
            self._refresh_list()
        self._new_var.set("")

    def _del_item(self):
        g   = self._group_var.get()
        sel = self._lb.curselection()
        if not sel or not g:
            return
        item = self._lb.get(sel[0])
        if item in self._data[g]:
            self._data[g].remove(item)
            self._refresh_list()

    def _save(self):
        save_ingredients(self._data)
        self._on_save(self._data)
        self.destroy()

# ─── Validation Panel ────────────────────────────────────────────────────────

class ValidationPanel(tk.Frame):
    """Shows warnings: unknown ingredients, missing/non-numeric usage."""

    def __init__(self, parent):
        super().__init__(parent, bg=PANEL_BG)
        self._ingr_master = []
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=PANEL_BG)
        hdr.pack(fill="x", padx=10, pady=(10, 2))
        tk.Label(hdr, text="⚠  Validation Warnings",
                 font=("TkDefaultFont", 9, "bold"),
                 bg=PANEL_BG, fg="#9B6E25").pack(side="left")
        self._count_lbl = tk.Label(hdr, text="",
                                   font=("TkDefaultFont", 8),
                                   bg=PANEL_BG, fg=TEXT_MUTED)
        self._count_lbl.pack(side="right")

        txt_frame = tk.Frame(self, bg=PANEL_BG)
        txt_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        vsb = tk.Scrollbar(txt_frame)
        vsb.pack(side="right", fill="y")

        self._txt = tk.Text(txt_frame, font=("TkDefaultFont", 8),
                            height=5,  #
                            bg="#FFFBF4", fg=TEXT_MAIN,
                            wrap="word", state="disabled",
                            bd=1, relief="solid",
                            yscrollcommand=vsb.set,
                            padx=6, pady=4)
        self._txt.pack(fill="both", expand=True)
        vsb.config(command=self._txt.yview)

        self._txt.tag_configure("sheet", foreground=ACCENT,  font=("TkDefaultFont", 8, "bold"))
        self._txt.tag_configure("warn",  foreground="#B03A2E")
        self._txt.tag_configure("ok",    foreground=BTN_ADD, font=("TkDefaultFont", 8, "bold"))
        self._txt.tag_configure("muted", foreground=TEXT_MUTED)

    def update_master(self, ingr_master):
        self._ingr_master = ingr_master

    def run(self, recipes):
        self._txt.config(state="normal")
        self._txt.delete("1.0", "end")

        master_set   = set(self._ingr_master)
        total_warns  = 0

        for rd in recipes:
            warns = []

            # Unknown ingredients
            for ingr_name in rd.ingredients:
                if ingr_name and ingr_name not in master_set:
                    warns.append(f'  \u2022 Not in master list: "{ingr_name}"')

            # Missing / non-numeric usage
            n = len(rd.ingredients)
            for row in rd.rows:
                for i in range(1, n + 1):
                    val = row.get(f"Usage{i}", "")
                    ing = rd.ingredients[i-1] if i-1 < len(rd.ingredients) else f"Ingredient{i}"
                    sz  = row.get("Size", "?")
                    op  = row.get("Option", "?")
                    sg  = row.get("Sugar Level", "?")
                    if val is None or str(val).strip() == "":
                        warns.append(f"  \u2022 Missing usage: {ing} [{sz}/{op}/{sg}]")
                    elif not is_numeric(val):
                        warns.append(f'  \u2022 Non-numeric usage "{val}": {ing} [{sz}/{op}/{sg}]')

            if warns:
                total_warns += len(warns)
                self._txt.insert("end", f"[{rd.name}]\n", "sheet")
                for w in warns:
                    self._txt.insert("end", w + "\n", "warn")
                self._txt.insert("end", "\n", "muted")

        if total_warns == 0:
            self._txt.insert("end", "✓  No issues found.", "ok")
            self._count_lbl.config(text="All OK", fg=BTN_ADD)
        else:
            self._count_lbl.config(text=f"{total_warns} warning(s)", fg="#B03A2E")

        self._txt.config(state="disabled")
        self._txt.see("1.0")

# ─── Sidebar ──────────────────────────────────────────────────────────────────

class Sidebar(tk.Frame):
    def __init__(self, parent, on_select, on_add, on_delete, on_duplicate, on_reorder=None):
        super().__init__(parent, bg=SIDEBAR_BG, width=220)
        self.pack_propagate(False)
        self.on_select    = on_select
        self.on_add       = on_add
        self.on_delete    = on_delete
        self.on_duplicate = on_duplicate
        self.on_reorder   = on_reorder   # Fix 1: drag reorder callback
        self._selected    = None         # primary (anchor) selection index
        self._multi_sel   = set()        # set of currently-selected indices
        self._sel_anchor  = None         # anchor for Shift+click range selection
        self._drag_idx    = None
        self._names       = []
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=SIDEBAR_BG, pady=20)
        top.pack(fill="x")
        tk.Label(top, text="🍵", font=("TkDefaultFont", 28),
                 bg=SIDEBAR_BG, fg=ACCENT).pack()
        tk.Label(top, text="Tea Recipes", font=("Georgia", 13, "bold"),
                 bg=SIDEBAR_BG, fg=TEXT_LIGHT).pack()
        tk.Label(top, text="Recipe Manager", font=("TkDefaultFont", 8),
                 bg=SIDEBAR_BG, fg=TEXT_MUTED).pack()

        tk.Frame(self, bg=ACCENT, height=1).pack(fill="x", padx=16, pady=(0, 12))

        bf = tk.Frame(self, bg=SIDEBAR_BG)
        bf.pack(fill="x", padx=12, pady=(0, 8))
        self._mk(bf, "＋ Add", BTN_ADD, self.on_add).pack(side="left", fill="x", expand=True, padx=(0, 2))
        self._mk(bf, "⧉", "#4A6FA5", self._dup_sel, width=3).pack(side="left", padx=2)
        self._mk(bf, "✕", BTN_DEL,   self._del_sel, width=3).pack(side="left")

        tk.Label(self, text="RECIPES", font=("TkDefaultFont", 8, "bold"),
                 bg=SIDEBAR_BG, fg=TEXT_MUTED, anchor="w", padx=16).pack(fill="x", pady=(4, 4))

        lf = tk.Frame(self, bg=SIDEBAR_BG)
        lf.pack(fill="both", expand=True)
        sc = tk.Scrollbar(lf, bg=SIDEBAR_BG, troughcolor=SIDEBAR_BG, bd=0)
        sc.pack(side="right", fill="y")
        self._cv = tk.Canvas(lf, bg=SIDEBAR_BG, highlightthickness=0, yscrollcommand=sc.set)
        self._cv.pack(fill="both", expand=True)
        sc.config(command=self._cv.yview)
        self._inn = tk.Frame(self._cv, bg=SIDEBAR_BG)
        self._win = self._cv.create_window((0, 0), window=self._inn, anchor="nw")
        self._inn.bind("<Configure>", lambda e: self._cv.configure(scrollregion=self._cv.bbox("all")))
        self._cv.bind("<Configure>", lambda e: self._cv.itemconfig(self._win, width=e.width))
        self._cv.bind("<MouseWheel>", lambda e: self._cv.yview_scroll(int(-1 * e.delta / 120), "units"))
        self._cv.bind("<Button-4>",   lambda e: self._cv.yview_scroll(-1, "units"))
        self._cv.bind("<Button-5>",   lambda e: self._cv.yview_scroll(1, "units"))

    def _mk(self, parent, text, color, cmd, width=None):
        kw = dict(text=text, font=("TkDefaultFont", 9, "bold"), fg="#FFF",
                  bg=color, bd=0, padx=8, pady=6, cursor="hand2", command=cmd, relief="flat")
        if width:
            kw["width"] = width
        return tk.Button(parent, **kw)

    def _selected_indices(self):
        """Return the current selection as a sorted list of indices."""
        if self._multi_sel:
            return sorted(self._multi_sel)
        if self._selected is not None:
            return [self._selected]
        return []

    def _del_sel(self):
        idxs = self._selected_indices()
        if not idxs:
            return
        # Pass list when multi, single int for backward compat
        if len(idxs) == 1:
            self.on_delete(idxs[0])
        else:
            self.on_delete(idxs)

    def _dup_sel(self):
        idxs = self._selected_indices()
        if not idxs:
            return
        if len(idxs) == 1:
            self.on_duplicate(idxs[0])
        else:
            self.on_duplicate(idxs)

    def populate(self, names, selected_idx=None, multi_sel=None):
        for w in self._inn.winfo_children():
            w.destroy()
        self._selected       = selected_idx
        self._names          = list(names)
        # Preserve multi-selection across re-populate when caller passes it,
        # otherwise reset to just the primary selection.
        if multi_sel is not None:
            self._multi_sel = set(i for i in multi_sel if 0 <= i < len(names))
        else:
            # Drop any stale indices that no longer exist
            self._multi_sel = {i for i in self._multi_sel if 0 <= i < len(names)}
            if not self._multi_sel and selected_idx is not None:
                self._multi_sel = {selected_idx}
        if self._sel_anchor is not None and self._sel_anchor >= len(names):
            self._sel_anchor = selected_idx
        self._drag_idx       = None   # index being dragged (None until threshold)
        self._drag_start_idx = None   # index where press started
        self._drag_start_y   = None   # y_root at press
        self._drop_target    = None   # current drop target index (live during drag)
        self._item_frames    = []     # frames for each row, used to find drop target
        self._drop_indicator = None   # visual line indicator widget

        def _scroll(e):
            self._cv.yview_scroll(int(-1 * e.delta / 120), "units")
        def _scroll_up(e):
            self._cv.yview_scroll(-1, "units")
        def _scroll_dn(e):
            self._cv.yview_scroll(1, "units")

        # Drop indicator line (a thin highlighted frame shown between items while dragging)
        self._drop_indicator = tk.Frame(self._inn, bg=ACCENT, height=3)

        # Tk modifier-state masks
        SHIFT_MASK = 0x0001
        CTRL_MASK  = 0x0004

        for i, name in enumerate(names):
            is_primary = (i == selected_idx)
            is_in_sel  = (i in self._multi_sel)
            # Primary selection: gold ACCENT; secondary multi-select: lighter highlight
            if is_primary:
                bg = ACCENT
                fg = "#1C1A17"
                handle_fg = "#1C1A17"
            elif is_in_sel:
                bg = MULTI_SEL_BG
                fg = TEXT_LIGHT
                handle_fg = "#1C1A17"
            else:
                bg = SIDEBAR_BG
                fg = TEXT_LIGHT
                handle_fg = TEXT_MUTED
            f  = tk.Frame(self._inn, bg=bg, cursor="hand2")
            f.pack(fill="x", padx=8, pady=2)
            self._item_frames.append(f)

            # Drag handle indicator
            tk.Label(f, text="⠿", font=("TkDefaultFont", 9),
                     bg=bg, fg=handle_fg,
                     padx=4).pack(side="left")

            prefix = "▶ " if is_primary else ("• " if is_in_sel else "  ")
            lbl = tk.Label(f, text=prefix + name,
                           font=("TkDefaultFont", 10), bg=bg, fg=fg,
                           anchor="w", padx=6, pady=8)
            lbl.pack(side="left", fill="x", expand=True)

            # ── Drag / click with modifier support ────────────────────
            # Strategy: track drop target on motion, but commit reorder only on release.
            # On release without movement: treat as click — apply Ctrl/Shift modifiers.
            def on_drag_start(e, idx=i):
                self._drag_idx   = None   # not committed yet
                self._drag_start_idx = idx
                self._drag_start_y   = e.y_root
                self._drag_start_state = e.state
                # If user starts dragging a row that's already part of the
                # multi-selection, drag the whole group. Otherwise drag only this row.
                if idx in self._multi_sel and len(self._multi_sel) > 1:
                    self._drag_group = sorted(self._multi_sel)
                else:
                    self._drag_group = [idx]
                self._drop_target    = idx

            def on_drag_motion(e, idx=i):
                # Only start dragging after moving more than 4px (distinguishes click from drag)
                if self._drag_start_idx is None:
                    return
                if self._drag_idx is None:
                    if abs(e.y_root - self._drag_start_y) < 4:
                        return
                    self._drag_idx = self._drag_start_idx
                # Find which item we're hovering over (or the gap above/below)
                widget = e.widget
                root_y = widget.winfo_rooty() + e.y
                target = self._find_drop_target(root_y)
                if target is not None and target != self._drop_target:
                    self._drop_target = target
                    self._show_drop_indicator(target)

            def on_drag_end(e, idx=i):
                # If mouse never moved enough → treat as a click (apply modifiers)
                if self._drag_idx is None and self._drag_start_idx == idx:
                    state = getattr(self, "_drag_start_state", 0) or 0
                    self._hide_drop_indicator()
                    self._drag_idx       = None
                    self._drag_start_idx = None
                    self._drag_start_y   = None
                    self._drop_target    = None
                    self._drag_group     = None
                    if state & SHIFT_MASK:
                        self._handle_shift_click(idx)
                    elif state & CTRL_MASK:
                        self._handle_ctrl_click(idx)
                    else:
                        self._handle_plain_click(idx)
                    return
                # Commit a reorder. If a group was being dragged, move all of them.
                group = getattr(self, "_drag_group", None) or [self._drag_start_idx]
                dst   = self._drop_target
                self._hide_drop_indicator()
                self._drag_idx       = None
                self._drag_start_idx = None
                self._drag_start_y   = None
                self._drop_target    = None
                self._drag_group     = None
                if dst is None:
                    return
                if len(group) == 1:
                    if group[0] != dst:
                        self._do_reorder(group[0], dst)
                else:
                    self._do_reorder_group(group, dst)

            for w in (f, lbl):
                w.bind("<ButtonPress-1>",   on_drag_start)
                w.bind("<B1-Motion>",       on_drag_motion)
                w.bind("<ButtonRelease-1>", on_drag_end)
                w.bind("<MouseWheel>", _scroll)
                w.bind("<Button-4>",   _scroll_up)
                w.bind("<Button-5>",   _scroll_dn)

            if not is_primary and not is_in_sel:
                def _enter(e, fr=f, lb=lbl):
                    fr.config(bg=BTN_HOVER); lb.config(bg=BTN_HOVER, fg=TEXT_MAIN)
                def _leave(e, fr=f, lb=lbl):
                    fr.config(bg=SIDEBAR_BG); lb.config(bg=SIDEBAR_BG, fg=TEXT_LIGHT)
                for w in (f, lbl):
                    w.bind("<Enter>", _enter)
                    w.bind("<Leave>", _leave)

    def _find_item_at_y(self, root_y):
        """Return index of sidebar item under the given root y coordinate."""
        for i, child in enumerate(self._item_frames):
            cy = child.winfo_rooty()
            ch = child.winfo_height()
            if cy <= root_y <= cy + ch:
                return i
        return None

    def _find_drop_target(self, root_y):
        """Return the index where the dragged item should be dropped.

        Uses the midpoint of each row: if the cursor is in the top half,
        drop *before* that row; bottom half, drop *after*. Clamps to ends
        so dragging above the first / below the last works."""
        if not self._item_frames:
            return None
        # Above the first item → drop at top
        first = self._item_frames[0]
        if root_y < first.winfo_rooty():
            return 0
        # Below the last item → drop at end
        last = self._item_frames[-1]
        if root_y > last.winfo_rooty() + last.winfo_height():
            return len(self._item_frames) - 1
        for i, child in enumerate(self._item_frames):
            cy = child.winfo_rooty()
            ch = child.winfo_height()
            if cy <= root_y <= cy + ch:
                return i
        return None

    def _show_drop_indicator(self, target_idx):
        """Show a highlight line at the prospective drop location."""
        if self._drop_indicator is None:
            return
        if target_idx is None or target_idx < 0 or target_idx >= len(self._item_frames):
            self._hide_drop_indicator()
            return
        # Place the indicator just above the target row
        try:
            self._drop_indicator.pack_forget()
        except Exception:
            pass
        try:
            self._drop_indicator.pack(fill="x", padx=8,
                                      before=self._item_frames[target_idx])
        except Exception:
            pass

    def _hide_drop_indicator(self):
        if self._drop_indicator is not None:
            try:
                self._drop_indicator.pack_forget()
            except Exception:
                pass

    def _do_reorder(self, from_idx, to_idx):
        """Reorder recipes by dragging from_idx to to_idx."""
        if self.on_reorder:
            self.on_reorder(from_idx, to_idx)

    def _do_reorder_group(self, group, to_idx):
        """Move a group of indices (sorted) so they land at to_idx, in order."""
        if self.on_reorder:
            self.on_reorder(group, to_idx)

    # ── Selection click handlers ─────────────────────────────────────────
    def _handle_plain_click(self, idx):
        """Plain click: select a single recipe (clear multi-selection)."""
        self._multi_sel = {idx}
        self._sel_anchor = idx
        self._selected = idx
        # Delegate to App: it will switch the active recipe and refresh sidebar
        self.on_select(idx)

    def _handle_ctrl_click(self, idx):
        """Ctrl+click: toggle this row's membership in the multi-selection."""
        if idx in self._multi_sel:
            self._multi_sel.discard(idx)
            # If we just removed the primary selection, pick another
            if idx == self._selected:
                if self._multi_sel:
                    new_primary = max(self._multi_sel) if idx > min(self._multi_sel) else min(self._multi_sel)
                    self._selected = new_primary
                    self._sel_anchor = new_primary
                    self.on_select(new_primary)
                    return
                else:
                    # Nothing selected anymore — fall through to repopulate
                    self._selected = None
                    self._sel_anchor = None
        else:
            self._multi_sel.add(idx)
            self._selected = idx
            self._sel_anchor = idx
            self.on_select(idx)
            return
        # Repopulate to reflect deselection
        self.populate(self._names, self._selected, self._multi_sel)

    def _handle_shift_click(self, idx):
        """Shift+click: select the range from anchor to clicked index."""
        anchor = self._sel_anchor if self._sel_anchor is not None else (
            self._selected if self._selected is not None else idx
        )
        lo, hi = (anchor, idx) if anchor <= idx else (idx, anchor)
        self._multi_sel = set(range(lo, hi + 1))
        self._selected = idx
        # Note: do not overwrite _sel_anchor — subsequent Shift+clicks extend from same anchor
        self.on_select(idx)

    def select_all(self):
        """Ctrl+A style: select every recipe."""
        if not self._names:
            return
        self._multi_sel = set(range(len(self._names)))
        self._sel_anchor = 0
        self._selected = self._selected if self._selected is not None else 0
        self.populate(self._names, self._selected, self._multi_sel)

    def clear_multi_selection(self):
        """Reduce selection back to just the primary index."""
        if self._selected is not None:
            self._multi_sel = {self._selected}
        else:
            self._multi_sel = set()
        self._sel_anchor = self._selected
        self.populate(self._names, self._selected, self._multi_sel)

# ─── Info Panel ───────────────────────────────────────────────────────────────

class InfoPanel(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=PANEL_BG)
        self.vars         = {}
        self._cbs         = {}
        self._tags_groups = copy.deepcopy(DEFAULT_TAGS_GROUPS)
        self._build()

    def _build(self):
        tk.Label(self, text="Recipe Information", font=("Georgia", 12, "bold"),
                 bg=PANEL_BG, fg=TEXT_MAIN).grid(
            row=0, column=0, columnspan=20, sticky="w", pady=(12, 8), padx=16)

        plain_fields = [
            ("Recipe Code",  "recipe_code",  12),
            ("Recipe Name",  "name",         24),
            ("Product Code", "product_code", 12),
        ]
        for ci, (label, key, w) in enumerate(plain_fields):
            col = ci * 2
            px  = (16 if ci == 0 else 8, 4)
            tk.Label(self, text=label, font=("TkDefaultFont", 8, "bold"),
                     bg=PANEL_BG, fg=TEXT_MUTED).grid(row=1, column=col, sticky="w", padx=px)
            v = tk.StringVar()
            self.vars[key] = v
            tk.Entry(self, textvariable=v, width=w,
                     font=("TkDefaultFont", 10), bd=1, relief="solid",
                     bg=BG, fg=TEXT_MAIN, insertbackground=TEXT_MAIN).grid(
                row=2, column=col, sticky="ew", padx=px, pady=(0, 12))

        # Tags dropdown (select country first)
        col = 6
        tk.Label(self, text="Tags", font=("TkDefaultFont", 8, "bold"),
                 bg=PANEL_BG, fg=TEXT_MUTED).grid(row=1, column=col, sticky="w", padx=(8, 4))
        v_tags = tk.StringVar()
        self.vars["tags"] = v_tags
        cb_tags = ttk.Combobox(self, textvariable=v_tags, width=16,
                                font=("TkDefaultFont", 10),
                                state="readonly",
                                values=list(self._tags_groups.keys()))
        cb_tags.grid(row=2, column=col, sticky="ew", padx=(8, 4), pady=(0, 12))
        self._cbs["tags"] = cb_tags

        # Group dropdown (filtered by selected tag)
        col = 8
        tk.Label(self, text="Group", font=("TkDefaultFont", 8, "bold"),
                 bg=PANEL_BG, fg=TEXT_MUTED).grid(row=1, column=col, sticky="w", padx=(8, 4))
        v_group = tk.StringVar()
        self.vars["group"] = v_group
        cb_group = ttk.Combobox(self, textvariable=v_group, width=18,
                                 font=("TkDefaultFont", 10), state="readonly", values=[])
        cb_group.grid(row=2, column=col, sticky="ew", padx=(8, 4), pady=(0, 12))
        self._cbs["group"] = cb_group

        # When tag changes → update group list
        def on_tag_change(event=None):
            tag = v_tags.get()
            groups = self._tags_groups.get(tag, [])
            cb_group["values"] = groups
            if v_group.get() not in groups:
                v_group.set(groups[0] if groups else "")
        cb_tags.bind("<<ComboboxSelected>>", on_tag_change)

    def update_groups_tags(self, tags_groups):
        self._tags_groups = tags_groups
        self._cbs["tags"]["values"] = list(tags_groups.keys())
        # Refresh group list for current tag
        tag = self.vars["tags"].get()
        groups = tags_groups.get(tag, [])
        self._cbs["group"]["values"] = groups

    def load(self, rd):
        self.vars["recipe_code"].set(rd.recipe_code or "")
        self.vars["name"].set(rd.name or "")
        self.vars["product_code"].set(rd.product_code or "")
        self.vars["tags"].set(rd.tags or "")
        # Refresh group dropdown to match the loaded tag
        tag = rd.tags or ""
        groups = self._tags_groups.get(tag, [])
        self._cbs["group"]["values"] = groups
        self.vars["group"].set(rd.group or "")

    def save(self, rd):
        rd.recipe_code  = self.vars["recipe_code"].get()
        rd.name         = self.vars["name"].get()
        rd.product_code = self.vars["product_code"].get()
        rd.group        = self.vars["group"].get()
        rd.tags         = self.vars["tags"].get()

# ─── Options Panel ────────────────────────────────────────────────────────────

class OptionsPanel(tk.Frame):
    def __init__(self, parent, on_change=None):
        super().__init__(parent, bg=PANEL_BG)
        self._on_change        = on_change
        self._size_option_vars = {}
        self._sugar_vars       = {}   # ordered: sugar_str -> BooleanVar
        self._custom_var       = tk.StringVar()
        self._suppress         = False
        self._build()

    def _build(self):
        # Size × Option matrix
        so = tk.Frame(self, bg=PANEL_BG)
        so.grid(row=0, column=0, sticky="nsew", padx=16, pady=(8, 4))

        tk.Label(so, text="Size & Ice Level Options",
                 font=("TkDefaultFont", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MUTED).grid(
            row=0, column=0, columnspan=20, sticky="w", pady=(4, 2))

        tk.Label(so, text="", bg=PANEL_BG, width=12).grid(row=1, column=0)
        for ci, opt in enumerate(OPTIONS, 1):
            tk.Label(so, text=opt,
                     font=("TkDefaultFont", 9, "bold"),
                     bg=PANEL_BG, fg=TEXT_MUTED,
                     width=13, anchor="center",      # ③ fixed wider width
                     wraplength=90).grid(             # ③ allow wrap
                row=1, column=ci, padx=4, pady=4)

        for ri, size in enumerate(SIZES, 2):
            lbl = f"M ({size})" if size == "16oz" else f"L ({size})"
            tk.Label(so, text=lbl, font=("TkDefaultFont", 10, "bold"),
                     bg=PANEL_BG, fg=TEXT_MAIN, anchor="w", width=10).grid(
                row=ri, column=0, padx=(8, 4), sticky="w")
            for ci, opt in enumerate(OPTIONS, 1):
                v = tk.BooleanVar(value=False)   # ① default unchecked
                self._size_option_vars[(size, opt)] = v
                tk.Checkbutton(so, variable=v, bg=PANEL_BG,
                               activebackground=PANEL_BG,
                               selectcolor=ACCENT,
                               command=self._changed).grid(row=ri, column=ci, padx=4)

        # Sugar level row (dynamic)
        self._sugar_outer = tk.Frame(self, bg=PANEL_BG)
        self._sugar_outer.grid(row=1, column=0, sticky="nsew", padx=16, pady=(4, 8))

        tk.Label(self._sugar_outer, text="Sugar Levels",
                 font=("TkDefaultFont", 9, "bold"),
                 bg=PANEL_BG, fg=TEXT_MUTED).pack(anchor="w", pady=(4, 2))

        self._sugar_inner = tk.Frame(self._sugar_outer, bg=PANEL_BG)
        self._sugar_inner.pack(fill="x")

        self._build_sugar_checks(STD_SUGARS)

    def _build_sugar_checks(self, sugars):
        for w in self._sugar_inner.winfo_children():
            w.destroy()
        old_vals = {s: v.get() for s, v in self._sugar_vars.items()}
        self._sugar_vars = {}

        ordered = sorted(set(sugars), key=sugar_key)
        for ci, sg in enumerate(ordered):
            v = tk.BooleanVar(value=old_vals.get(sg, False))
            self._sugar_vars[sg] = v
            f = tk.Frame(self._sugar_inner, bg=PANEL_BG)
            f.grid(row=0, column=ci, padx=4, pady=4)
            tk.Checkbutton(f, variable=v, bg=PANEL_BG,
                           activebackground=PANEL_BG,
                           selectcolor=ACCENT,
                           command=self._changed).pack(side="left")
            tk.Label(f, text=sg, font=("TkDefaultFont", 9),
                     bg=PANEL_BG, fg=TEXT_MAIN).pack(side="left")
            # Fix 7: show small ✕ only for custom (non-standard) sugar levels
            if sg not in STD_SUGARS:
                tk.Button(f, text="✕", font=("TkDefaultFont", 7),
                          fg=BTN_DEL, bg=PANEL_BG, bd=0,
                          padx=1, pady=0, cursor="hand2",
                          command=lambda s=sg: self._del_custom_sugar(s)).pack(side="left")

        # Custom entry
        cf = tk.Frame(self._sugar_inner, bg=PANEL_BG)
        cf.grid(row=0, column=len(ordered), padx=(12, 8), pady=4)
        tk.Label(cf, text="Custom:", font=("TkDefaultFont", 9),
                 bg=PANEL_BG, fg=TEXT_MUTED).pack(side="left")
        e = tk.Entry(cf, textvariable=self._custom_var, width=7,
                     font=("TkDefaultFont", 9), bd=1, relief="solid", bg=BG)
        e.pack(side="left", padx=4)
        e.bind("<Return>", self._add_custom)
        tk.Button(cf, text="＋", font=("TkDefaultFont", 8),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=4, pady=2,
                  cursor="hand2", command=self._add_custom).pack(side="left")

    def _del_custom_sugar(self, sg):
        """Remove a custom (non-standard) sugar level checkbox."""
        current = [s for s in self._sugar_vars.keys() if s != sg]
        self._build_sugar_checks(current)
        self._changed()

    def _add_custom(self, event=None):
        val = self._custom_var.get().strip()
        if not val:
            return
        if not val.endswith("%"):
            val += "%"
        current = list(self._sugar_vars.keys())
        if val not in current:
            current.append(val)
            self._build_sugar_checks(current)
            if val in self._sugar_vars:
                self._sugar_vars[val].set(True)
        self._custom_var.set("")
        self._changed()


    def _changed(self):
        if not self._suppress and self._on_change:
            self._on_change()

    def load(self, rd):
        self._suppress = True
        for v in self._size_option_vars.values():
            v.set(False)
        present_so    = set()
        present_sugar = set()
        for row in rd.rows:
            present_so.add((row.get("Size", ""), row.get("Option", "")))
            sg = str(row.get("Sugar Level", ""))
            if sg:
                present_sugar.add(sg)

        # ถ้าเป็นสูตรใหม่ (ยังไม่มีข้อมูล) ให้ติ๊ก Sugar พื้นฐานรอไว้เลย
        if not rd.rows:
            present_sugar = set(STD_SUGARS)

        # ⑤ Dynamic sugar checkboxes from file
        all_sugars = sorted(set(list(STD_SUGARS) + list(present_sugar)) - {""}, key=sugar_key)
        self._build_sugar_checks(all_sugars)

        for (size, opt), var in self._size_option_vars.items():
            var.set((size, opt) in present_so)
        for sg, var in self._sugar_vars.items():
            var.set(sg in present_sugar)
        self._suppress = False

    def get_active_combos(self):
        active_so = [(sz, op) for (sz, op), v in self._size_option_vars.items() if v.get()]
        sugars    = sorted([sg for sg, v in self._sugar_vars.items() if v.get()], key=sugar_key)
        combos = []
        for size in SIZES:
            for opt in OPTIONS:
                if (size, opt) in active_so:
                    for sg in sugars:
                        combos.append((size, opt, sg))
        return combos

    def get_active_options(self):
        return [op for op in OPTIONS
                if any(v.get() for (sz, op2), v in self._size_option_vars.items() if op2 == op)]

    def get_active_options_for_size(self, size):
        """Return options ticked specifically for the given size."""
        return [op for op in OPTIONS
                if self._size_option_vars.get((size, op), tk.BooleanVar()).get()]

    def get_active_sugars(self):
        return sorted([sg for sg, v in self._sugar_vars.items() if v.get()], key=sugar_key)

# ─── Tab Table (M / L tabs) ──────────────────────────────────────────────────

class TabTable(tk.Frame):
    """
    Two-tab view: Tab M (16oz) and Tab L (22oz).
    Each tab shows a pivot table:
      Rows = Ingredients (base or sugar-dependent)
      Cols = active Options
    Base rows: 1 row, fill all sugar levels on write.
    Sugar rows: 1 row per sugar level.
    Export still writes full flat format.
    """
    MAX_INGR = 8

    def __init__(self, parent, ingr_master, on_data_change=None, on_auto_sugar=None):
        super().__init__(parent, bg=BG)
        self._rd             = None
        self._ingr_master    = ingr_master
        self._on_data_change = on_data_change
        self._on_auto_sugar  = on_auto_sugar   # Fix 6: callback to auto-check sugars
        self._active_opts_by_size = {sz: OPTIONS[:] for sz in SIZES}
        self._active_sugars       = STD_SUGARS[:]
        self._ingr_labels    = []   # actual ingredient names (display + Excel value)
        self._ingr_is_sugar  = []   # bool: True = sugar-dependent
        self._build()

    # ── Build chrome (header + notebook) ─────────────────────────────────

    def _build(self):
        # Header bar
        h = tk.Frame(self, bg=PANEL_BG)
        h.pack(fill="x")
        tk.Label(h, text="Ingredient Usage Table",
                 font=("Georgia", 12, "bold"),
                 bg=PANEL_BG, fg=TEXT_MAIN).pack(side="left", padx=16, pady=10)

        right = tk.Frame(h, bg=PANEL_BG)
        right.pack(side="right", padx=12, pady=8)
        tk.Button(right, text="＋ Ingredient", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_ADD, bd=0, padx=8, pady=4,
                  cursor="hand2", command=self._add_ingredient).pack(side="left", padx=2)
        tk.Button(right, text="－ Remove Last", font=("TkDefaultFont", 9),
                  fg="#FFF", bg=BTN_DEL, bd=0, padx=8, pady=4,
                  cursor="hand2", command=self._remove_ingredient).pack(side="left", padx=2)

        # ttk Notebook for M / L tabs
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Tea.TNotebook", background=BG, borderwidth=0)
        style.configure("Tea.TNotebook.Tab",
                        font=("TkDefaultFont", 10, "bold"),
                        padding=(16, 6),
                        background=BORDER_CLR,
                        foreground=TEXT_MUTED)
        style.map("Tea.TNotebook.Tab",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", "#1C1A17")])

        self._nb = ttk.Notebook(self, style="Tea.TNotebook")
        self._nb.pack(fill="both", expand=True, padx=16, pady=(4, 16))

        self._tab_frames = {}
        tab_labels = {"16oz": "M  (16 oz)", "22oz": "L  (22 oz)"}
        for size in SIZES:
            f = tk.Frame(self._nb, bg=BG)
            self._nb.add(f, text=tab_labels[size])
            self._tab_frames[size] = f

    # ── Load data ─────────────────────────────────────────────────────────

    def load(self, rd, active_options, active_sugars):
        self._rd  = rd
        # active_options: dict {size: [options]} OR flat list (legacy)
        if isinstance(active_options, dict):
            self._active_opts_by_size = active_options
        else:
            opts = active_options if active_options else OPTIONS[:]
            self._active_opts_by_size = {sz: opts for sz in SIZES}
        self._active_sugars  = active_sugars if active_sugars else STD_SUGARS[:]
        self._ingr_labels    = rd.ingredients[:]
        self._ingr_is_sugar  = [
            self._guess_sugar_dep(rd, i + 1)
            for i in range(len(rd.ingredients))
        ]
        self._render_all()

    def _guess_sugar_dep(self, rd, ingr_idx):
        # Rule: name contains "Syrup" (case-insensitive) → always Sug.
        #       everything else → always Base.
        ingr_name = rd.ingredients[ingr_idx - 1] if ingr_idx - 1 < len(rd.ingredients) else ""
        return "syrup" in str(ingr_name).lower()

    def _render_all(self):
        for size in SIZES:
            self._render_tab(size)

    # ── Render one tab ────────────────────────────────────────────────────

    def _render_tab(self, size):
        frame = self._tab_frames[size]
        for w in frame.winfo_children():
            w.destroy()
        # Fix 1: close any stray dropdowns when tab re-renders

        # Scrollable canvas inside tab
        hsc = tk.Scrollbar(frame, orient="horizontal")
        hsc.pack(side="bottom", fill="x")
        vsc = tk.Scrollbar(frame)
        vsc.pack(side="right", fill="y")
        cv = tk.Canvas(frame, bg=BG, highlightthickness=0,
                       xscrollcommand=hsc.set, yscrollcommand=vsc.set)
        cv.pack(fill="both", expand=True)
        hsc.config(command=cv.xview)
        vsc.config(command=cv.yview)

        tbl = tk.Frame(cv, bg=BG)
        win = cv.create_window((0, 0), window=tbl, anchor="nw")
        tbl.bind("<Configure>", lambda e, cv=cv: cv.config(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>", lambda e, cv=cv, win=win, tbl=tbl:
                cv.itemconfig(win, width=max(e.width, tbl.winfo_reqwidth())))
        cv.bind("<MouseWheel>", lambda e, cv=cv:
                cv.yview_scroll(int(-1 * e.delta / 120), "units"))

        size_opts = self._active_opts_by_size.get(size, [])
        options   = [o for o in OPTIONS if o in size_opts]
        sugars    = self._active_sugars

        if not options or not sugars:
            tk.Label(tbl,
                     text="No options selected.\nTick checkboxes above to add rows.",
                     font=("TkDefaultFont", 10),
                     bg=BG, fg=TEXT_MUTED, pady=20, justify="center").grid(row=0, column=0)
            return

        # Build lookup: (size, option, sugar) -> row_dict
        lookup = {}
        for row in (self._rd.rows if self._rd else []):
            k = (row.get("Size"), row.get("Option"), row.get("Sugar Level"))
            lookup[k] = row

        # Column widths
        INGR_W = 22
        OPT_W  = 10
        SUG_W  = 7
        TYP_W  = 5

        # Header row
        col = 0
        self._hcell(tbl, 0, col, "Ingredient", INGR_W); col += 1
        self._hcell(tbl, 0, col, "Type",       TYP_W);  col += 1
        self._hcell(tbl, 0, col, "Sugar",      SUG_W);  col += 1
        for opt in options:
            self._hcell(tbl, 0, col, opt, OPT_W); col += 1
        self._hcell(tbl, 0, col, "Pop-up Tips", 16)

        # Data rows
        row_num = 1
        for ni in range(len(self._ingr_labels)):
            ingr_lbl  = self._ingr_labels[ni]
            ingr_idx  = ni + 1
            usage_key = f"Usage{ingr_idx}"
            is_sugar  = self._ingr_is_sugar[ni] if ni < len(self._ingr_is_sugar) else False

            if is_sugar and sugars:
                n_s = len(sugars)
                for si, sugar in enumerate(sugars):
                    bg  = PANEL_BG if row_num % 2 == 0 else ROW_ALT
                    col = 0
                    if si == 0:
                        self._ingr_cell(tbl, row_num, col, ni, ingr_lbl, bg, INGR_W, rowspan=n_s)
                    col += 1
                    if si == 0:
                        self._type_btn(tbl, row_num, col, ni, is_sugar, bg, size, rowspan=n_s)
                    col += 1
                    tk.Label(tbl, text=sugar, font=("TkDefaultFont", 9),
                             bg=bg, fg=TEXT_MUTED, width=SUG_W, pady=4).grid(
                        row=row_num, column=col, padx=1, sticky="nsew")
                    col += 1
                    for opt in options:
                        rd_ = lookup.get((size, opt, sugar), {})
                        fk = (size, opt, sugar) # เก็บค่า Size, Option, Sugar ปัจจุบัน
                        self._usage_entry(tbl, row_num, col, rd_, usage_key, bg, OPT_W,
                                          ingr_lbl=ingr_lbl, fk=fk, lookup=lookup)
                        col += 1
                    if si == 0:
                        fk = (size, options[0], sugars[0])
                        self._tips_entry(tbl, row_num, col,
                                         lookup.get(fk, {}), bg, 16, rowspan=n_s)
                    row_num += 1
            else:
                bg  = PANEL_BG if row_num % 2 == 0 else ROW_ALT
                col = 0
                self._ingr_cell(tbl, row_num, col, ni, ingr_lbl, bg, INGR_W)
                col += 1
                self._type_btn(tbl, row_num, col, ni, is_sugar, bg, size)
                col += 1
                tk.Label(tbl, text="(all)", font=("TkDefaultFont", 8),
                         bg=bg, fg=TEXT_MUTED, width=SUG_W, pady=4).grid(
                    row=row_num, column=col, padx=1, sticky="nsew")
                col += 1
                first_sg = sugars[0] if sugars else "0%"
                for opt in options:
                    all_keys = [(size, opt, sg) for sg in sugars]
                    rd_ = lookup.get((size, opt, first_sg), {})
                    self._usage_entry(tbl, row_num, col, rd_, usage_key, bg, OPT_W,
                                      base_fill=True, all_keys=all_keys, lookup=lookup)
                    col += 1
                fk = (size, options[0], first_sg) if options else None
                self._tips_entry(tbl, row_num, col, lookup.get(fk, {}) if fk else {}, bg, 16)
                row_num += 1

    # ── Cell helpers ──────────────────────────────────────────────────────

    def _hcell(self, t, row, col, text, w=12, rowspan=1):
        tk.Label(t, text=text, font=("TkDefaultFont", 9, "bold"),
                 bg=HEADER_BG, fg=TEXT_LIGHT, width=w, pady=6, padx=4,
                 wraplength=w * 8, anchor="center").grid(
            row=row, column=col, padx=1, pady=1, sticky="nsew", rowspan=rowspan)

    def _ingr_cell(self, t, row, col, ni, ingr_lbl, bg, w=22, rowspan=1):
        f = tk.Frame(t, bg=bg)
        f.grid(row=row, column=col, padx=1, pady=0, sticky="nsew", rowspan=rowspan)
        # Combobox takes most of the width; ✕ delete button sits at the right
        cb = SearchCombobox(f, self._ingr_master, width=max(4, w - 5), bg=bg,
                            on_select=lambda val, ni=ni: self._set_ingr(ni, val, render=True))
        cb.pack(side="left", fill="x", expand=True, padx=(2, 0), pady=2)
        cb.set(ingr_lbl)  # set() uses _suppress=True so trace does NOT fire here
        cb._var.trace_add("write", lambda *a, cb=cb, ni=ni: self._set_ingr(ni, cb.get(), render=False))
        # Up / Down move buttons
        n_total = len(self._ingr_labels)
        up_btn = tk.Button(f, text="▲", font=("TkDefaultFont", 7),
                           fg=TEXT_LIGHT, bg="#5A5450", bd=0, padx=3, pady=0,
                           cursor="hand2" if ni > 0 else "arrow",
                           state="normal" if ni > 0 else "disabled",
                           command=lambda ni=ni: self._move_ingredient(ni, ni - 1))
        up_btn.pack(side="right", padx=(0, 1), pady=2)
        dn_btn = tk.Button(f, text="▼", font=("TkDefaultFont", 7),
                           fg=TEXT_LIGHT, bg="#5A5450", bd=0, padx=3, pady=0,
                           cursor="hand2" if ni < n_total - 1 else "arrow",
                           state="normal" if ni < n_total - 1 else "disabled",
                           command=lambda ni=ni: self._move_ingredient(ni, ni + 1))
        dn_btn.pack(side="right", padx=(0, 1), pady=2)
        # Per-row delete button
        del_btn = tk.Button(f, text="✕", font=("TkDefaultFont", 8, "bold"),
                            fg="#FFF", bg=BTN_DEL, bd=0, padx=4, pady=0,
                            cursor="hand2",
                            command=lambda ni=ni: self._remove_ingredient_at(ni))
        del_btn.pack(side="right", padx=(2, 2), pady=2)

    def _set_ingr(self, ni, val, render=False):
        if ni < len(self._ingr_labels):
            self._ingr_labels[ni] = val
            
            # ฟิกซ์บั๊ก: บันทึกค่าที่เลือกลง object หลักทันที เพื่อไม่ให้โดนข้อมูลเก่าทับตอน Re-render
            if self._rd:
                self.collect(self._rd)

            key = f"Ingredient{ni + 1}"
            for row in (self._rd.rows if self._rd else []):
                row[key] = val
            # Rule: only ingredients whose name contains "Syrup" can be Sug.
            # Everything else is always Base.
            if ni < len(self._ingr_is_sugar):
                if "syrup" in str(val).lower():
                    self._ingr_is_sugar[ni] = True
                else:
                    self._ingr_is_sugar[ni] = False
            if render:
                # Only re-render when explicitly selected (not on every keystroke)
                # Auto-check sugars then re-render table to show Sug rows
                if val == SYRUP_LAST_NAME and self._on_auto_sugar:
                    self._on_auto_sugar()  # this triggers _render_all via chain
                else:
                    self._render_all()
            if self._on_data_change:
                self._on_data_change()

    def _type_btn(self, t, row, col, ni, is_sugar, bg, size, rowspan=1):
        lbl   = "Sug" if is_sugar else "Base"
        color = "#4A6FA5" if is_sugar else BTN_ADD

        def toggle(ni=ni):
            self._ingr_is_sugar[ni] = not self._ingr_is_sugar[ni]
            self._render_all()

        state = "normal"
        tk.Button(t, text=lbl, font=("TkDefaultFont", 7),
                  fg="#FFF", bg=color, bd=0, padx=2, pady=2,
                  cursor="hand2",
                  state=state, command=toggle).grid(
            row=row, column=col, padx=1, pady=1, sticky="nsew", rowspan=rowspan)

    def _usage_entry(self, t, row, col, row_dict, usage_key, bg, w=10,
                     base_fill=False, all_keys=None, lookup=None,
                     ingr_lbl=None, fk=None):
        val = row_dict.get(usage_key, "")
        var = tk.StringVar(value=str(val) if val is not None else "")
        
        # บันทึกตัวแปร UI ไว้เพื่อให้อัปเดตข้ามช่องได้อัตโนมัติ
        if fk:
            if not hasattr(self, '_usage_vars'):
                self._usage_vars = {}
            self._usage_vars[(fk[0], fk[1], fk[2], usage_key)] = var

        e   = tk.Entry(t, textvariable=var, width=w,
                       font=("TkDefaultFont", 9),
                       bg=bg, fg=TEXT_MAIN,
                       insertbackground=TEXT_MAIN,
                       bd=0, relief="flat", justify="center")
        e.grid(row=row, column=col, padx=1, pady=0, ipady=4, sticky="nsew")

        def on_enter(event, t=t, r=row, c=col):
            next_row = r + 1
            widget = t.grid_slaves(row=next_row, column=c)
            if widget:
                widget[0].focus_set()

        e.bind("<Return>", on_enter)

        def on_write(*_):
            ok = is_numeric(var.get())
            e.config(bg=bg if ok else ERR_BG, fg=TEXT_MAIN if ok else ERR_FG)
            row_dict[usage_key] = var.get()
            
            if base_fill and all_keys and lookup:
                for k in all_keys:
                    if k in lookup:
                        lookup[k][usage_key] = var.get()
                        
            # --- ลอจิก Auto-Fill สำหรับ Syrup (75%) ---
            if ingr_lbl == SYRUP_LAST_NAME and fk and fk[2] == "75%":
                try:
                    base_v = float(var.get())
                    updates = {
                        "100%": base_v + 5,
                        "50%":  3 if (base_v - 5) <= 0 else (base_v - 5),
                        "25%":  3 if (base_v - 10) <= 0 else (base_v - 10),
                        "0%":   0
                    }
                    for sg, nv in updates.items():
                        # ปัดเศษทิ้งถ้าเป็นจำนวนเต็ม
                        nv_str = str(int(nv)) if nv.is_integer() else str(nv)
                        target_fk = (fk[0], fk[1], sg)
                        
                        # 1. อัปเดตข้อมูลเบื้องหลัง
                        if lookup and target_fk in lookup:
                            lookup[target_fk][usage_key] = nv_str
                            
                        # 2. อัปเดตตัวเลขโชว์บนหน้าจอทันที
                        if hasattr(self, '_usage_vars'):
                            target_var = self._usage_vars.get((fk[0], fk[1], sg, usage_key))
                            if target_var and target_var.get() != nv_str:
                                target_var.set(nv_str)
                except ValueError:
                    pass # ปล่อยผ่านถ้าช่องว่างหรือพิมพ์ตัวอักษรผิด

            if self._on_data_change:
                self._on_data_change()

        var.trace_add("write", lambda *a: on_write())

    def _tips_entry(self, t, row, col, row_dict, bg, w=16, rowspan=1):
        val = row_dict.get(POPUP_HDR, "")
        var = tk.StringVar(value=str(val) if val is not None else "")
        # Fix 3: Pop-up Tips is read-only (grey, no editing)
        e   = tk.Entry(t, textvariable=var, width=w,
                       font=("TkDefaultFont", 9),
                       bg="#E8E8E8", fg="#AAAAAA",
                       disabledbackground="#E8E8E8", disabledforeground="#AAAAAA",
                       bd=0, relief="flat", state="disabled")
        e.grid(row=row, column=col, padx=1, pady=0, ipady=4, sticky="nsew", rowspan=rowspan)

    # ── Ingredient management ─────────────────────────────────────────────

    def _add_ingredient(self):
        if len(self._ingr_labels) >= self.MAX_INGR:
            messagebox.showinfo("Limit", f"Maximum {self.MAX_INGR} ingredients.")
            return
        # Insert before the first Sug (Syrup) ingredient, or at the end if none exists.
        insert_at = next(
            (i for i, lbl in enumerate(self._ingr_labels)
             if "syrup" in str(lbl).lower()),
            len(self._ingr_labels)
        )
        self._ingr_labels.insert(insert_at, "")
        self._ingr_is_sugar.insert(insert_at, False)
        # Shift Ingredient{N}/Usage{N} keys in all rows to match new order
        if self._rd:
            n = len(self._ingr_labels)  # after insert
            for row in self._rd.rows:
                # Shift existing keys from the end down to make room at insert_at
                for j in range(n - 1, insert_at, -1):
                    row[f"Ingredient{j + 1}"] = row.pop(f"Ingredient{j}", "")
                    row[f"Usage{j + 1}"]      = row.pop(f"Usage{j}", "")
                # Clear the new slot
                row[f"Ingredient{insert_at + 1}"] = ""
                row[f"Usage{insert_at + 1}"]      = ""
            self._rd.ingredients = self._ingr_labels[:]
        self._render_all()

    def _remove_ingredient(self):
        if not self._ingr_labels:
            return
        # "Remove Last" button: drop the last ingredient
        self._remove_ingredient_at(len(self._ingr_labels) - 1)

    def _remove_ingredient_at(self, ni):
        """Remove the ingredient at position ni (0-based) and re-number rows."""
        if ni is None or ni < 0 or ni >= len(self._ingr_labels):
            return
        # Confirm before removing a named ingredient
        name = self._ingr_labels[ni] if ni < len(self._ingr_labels) else ""
        if name:
            if not messagebox.askyesno(
                    "Remove Ingredient",
                    f"Remove ingredient '{name}'?\nUsage values for this column will be deleted."):
                return
        # Shift Ingredient{N}/Usage{N} keys down for all rows beyond ni
        n = len(self._ingr_labels)
        for row in (self._rd.rows if self._rd else []):
            # Drop the one being removed
            row.pop(f"Ingredient{ni + 1}", None)
            row.pop(f"Usage{ni + 1}", None)
            # Shift every higher index one slot down
            for j in range(ni + 1, n):
                src_i = f"Ingredient{j + 1}"
                src_u = f"Usage{j + 1}"
                dst_i = f"Ingredient{j}"
                dst_u = f"Usage{j}"
                if src_i in row:
                    row[dst_i] = row.pop(src_i)
                if src_u in row:
                    row[dst_u] = row.pop(src_u)
        # Remove from local lists
        self._ingr_labels.pop(ni)
        if ni < len(self._ingr_is_sugar):
            self._ingr_is_sugar.pop(ni)
        if self._rd:
            self._rd.ingredients = self._ingr_labels[:]
        self._render_all()
        if self._on_data_change:
            self._on_data_change()

    def _move_ingredient(self, from_ni, to_ni):
        """Swap ingredient at from_ni with to_ni (move up or down by one)."""
        n = len(self._ingr_labels)
        if from_ni < 0 or from_ni >= n or to_ni < 0 or to_ni >= n:
            return
        # Swap labels and type flags
        self._ingr_labels[from_ni], self._ingr_labels[to_ni] = \
            self._ingr_labels[to_ni], self._ingr_labels[from_ni]
        self._ingr_is_sugar[from_ni], self._ingr_is_sugar[to_ni] = \
            self._ingr_is_sugar[to_ni], self._ingr_is_sugar[from_ni]
        # Swap Ingredient{N}/Usage{N} keys in every row
        for row in (self._rd.rows if self._rd else []):
            ki_a = f"Ingredient{from_ni + 1}"
            ku_a = f"Usage{from_ni + 1}"
            ki_b = f"Ingredient{to_ni + 1}"
            ku_b = f"Usage{to_ni + 1}"
            row[ki_a], row[ki_b] = row.get(ki_b, ""), row.get(ki_a, "")
            row[ku_a], row[ku_b] = row.get(ku_b, ""), row.get(ku_a, "")
        if self._rd:
            self._rd.ingredients = self._ingr_labels[:]
        self._render_all()
        if self._on_data_change:
            self._on_data_change()

    def collect(self, rd):
        rd.ingredients = self._ingr_labels[:]

    def has_validation_errors(self):
        for row in (self._rd.rows if self._rd else []):
            for i in range(1, len(self._ingr_labels) + 1):
                if not is_numeric(row.get(f"Usage{i}", "")):
                    return True
        return False

    def update_ingr_master(self, items):
        self._ingr_master = items

# ─── Undo Manager ────────────────────────────────────────────────────────────

class UndoManager:
    def __init__(self, limit=UNDO_LIMIT):
        self._limit   = limit
        self._history = []
        self._pos     = -1

    def snapshot(self, recipes):
        snap = [rd.deep_copy() for rd in recipes]
        self._history = self._history[:self._pos + 1]
        self._history.append(snap)
        if len(self._history) > self._limit:
            self._history.pop(0)
        self._pos = len(self._history) - 1

    def can_undo(self): return self._pos > 0
    def can_redo(self): return self._pos < len(self._history) - 1

    def undo(self):
        if self.can_undo():
            self._pos -= 1
            return [rd.deep_copy() for rd in self._history[self._pos]]

    def redo(self):
        if self.can_redo():
            self._pos += 1
            return [rd.deep_copy() for rd in self._history[self._pos]]

# ─── Main App ────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tea Recipe Manager v2")
        self.geometry("1300x700")
        self.minsize(960, 640)
        self.configure(bg=BG)

        self._recipes   = []
        self._current   = None
        self._file_path = None
        self._ingr_data = load_ingredients()
        load_code_mappings()
        self._tags_groups = load_groups_tags()   # dict {tag: [groups]}
        self._undo_mgr  = UndoManager()
        self._row_cache = {}   # {recipe_name: {(size,option,sugar): row_dict}}
        self._dirty     = False  # True when unsaved changes exist
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build_ui()
        self._new_recipe_set()

    def _build_ui(self):
        # ── Menu ──────────────────────────────────────────────────────────
        mb = tk.Menu(self, bg="#2E2B25", fg=TEXT_LIGHT,
                     activebackground=ACCENT, activeforeground="#1C1A17")
        fm = tk.Menu(mb, tearoff=0, bg="#2E2B25", fg=TEXT_LIGHT,
                     activebackground=ACCENT, activeforeground="#1C1A17")
        fm.add_command(label="New",               command=self._new_recipe_set, accelerator="Ctrl+N")
        fm.add_command(label="Open...",           command=self._open_file,      accelerator="Ctrl+O")
        fm.add_command(label="Import from File...", command=self._import_file)
        fm.add_separator()
        fm.add_command(label="Save",              command=self._save_file,      accelerator="Ctrl+S")
        fm.add_command(label="Save As...",        command=self._save_as_file)
        fm.add_separator()
        fm.add_command(label="Exit",              command=self.quit)
        mb.add_cascade(label="File", menu=fm)

        tm = tk.Menu(mb, tearoff=0, bg="#2E2B25", fg=TEXT_LIGHT,
                     activebackground=ACCENT, activeforeground="#1C1A17")
        tm.add_command(label="Manage Ingredients...",   command=self._open_ingr_mgr)
        tm.add_command(label="Manage Code Mappings...", command=self._open_code_mgr)
        tm.add_command(label="Manage Groups & Tags...", command=self._open_groups_tags_mgr)
        mb.add_cascade(label="Tools", menu=tm)
        self.config(menu=mb)

        self.bind("<Control-n>", lambda e: self._new_recipe_set())
        self.bind("<Control-o>", lambda e: self._open_file())
        self.bind("<Control-s>", lambda e: self._save_file())
        self.bind("<Control-z>", lambda e: self._undo())
        self.bind("<Control-y>", lambda e: self._redo())
        # Sidebar selection shortcuts: Ctrl+A select all, Escape clears multi-select
        self.bind("<Control-a>", lambda e: self.sidebar.select_all())
        self.bind("<Control-A>", lambda e: self.sidebar.select_all())
        self.bind("<Escape>",    lambda e: self.sidebar.clear_multi_selection())
        # Delete key removes the current selection (single or group)
        self.bind("<Delete>",    lambda e: self.sidebar._del_sel())

        # ── Pane ──────────────────────────────────────────────────────────
        pane = tk.PanedWindow(self, orient="horizontal", bg=BG,
                              sashrelief="flat", sashwidth=4, sashpad=0, handlesize=0)
        pane.pack(fill="both", expand=True)

        self.sidebar = Sidebar(pane,
                               on_select=self._select_recipe,
                               on_add=self._add_recipe,
                               on_delete=self._delete_recipe,
                               on_duplicate=self._duplicate_recipe,
                               on_reorder=self._reorder_recipes)
        pane.add(self.sidebar, minsize=200, width=220)

        content = tk.Frame(pane, bg=BG)
        pane.add(content, minsize=640)

        # ── Toolbar ───────────────────────────────────────────────────────
        tb = tk.Frame(content, bg=HEADER_BG, pady=8, padx=16)
        tb.pack(fill="x")

        self._title_lbl = tk.Label(tb, text="New Recipe",
                                   font=("Georgia", 15, "bold"),
                                   bg=HEADER_BG, fg=TEXT_LIGHT, anchor="w")
        self._title_lbl.pack(side="left")

        self._file_lbl = tk.Label(tb, text="Unsaved",
                                  font=("TkDefaultFont", 9),
                                  bg=HEADER_BG, fg=TEXT_MUTED)
        self._file_lbl.pack(side="left", padx=24)

        # Right-side buttons  (pack order: rightmost first when using side="right")
        tk.Button(tb, text="📋  Export QR",
                  font=("TkDefaultFont", 10, "bold"),
                  fg="#FFFFFF", bg=BTN_ADD, bd=0, padx=16, pady=6,
                  cursor="hand2", command=self._export_qr).pack(side="right", padx=4)
        tk.Button(tb, text="💾  Save Excel",
                  font=("TkDefaultFont", 10, "bold"),
                  fg="#1C1A17", bg=ACCENT, bd=0, padx=16, pady=6,
                  cursor="hand2", command=self._save_file).pack(side="right", padx=4)
        tk.Button(tb, text="📂  Open",
                  font=("TkDefaultFont", 10),
                  fg=TEXT_LIGHT, bg=HEADER_BG, bd=1, relief="solid",
                  padx=10, pady=6, cursor="hand2",
                  command=self._open_file).pack(side="right", padx=4)


        self._redo_btn = tk.Button(tb, text="↪ Redo",
                                   font=("TkDefaultFont", 9),
                                   fg=TEXT_LIGHT, bg=HEADER_BG, bd=1, relief="solid",
                                   padx=8, pady=4, cursor="hand2",
                                   state="disabled", command=self._redo)
        self._redo_btn.pack(side="right", padx=2)

        self._undo_btn = tk.Button(tb, text="↩ Undo",
                                   font=("TkDefaultFont", 9),
                                   fg=TEXT_LIGHT, bg=HEADER_BG, bd=1, relief="solid",
                                   padx=8, pady=4, cursor="hand2",
                                   state="disabled", command=self._undo)
        self._undo_btn.pack(side="right", padx=2)

        # ── Scrollable main area ──────────────────────────────────────────
        sc = tk.Frame(content, bg=BG)
        sc.pack(fill="both", expand=True)
        vsb = tk.Scrollbar(sc)
        vsb.pack(side="right", fill="y")
        self._mc = tk.Canvas(sc, bg=BG, highlightthickness=0, yscrollcommand=vsb.set)
        self._mc.pack(fill="both", expand=True)
        vsb.config(command=self._mc.yview)
        self._mf = tk.Frame(self._mc, bg=BG)
        self._mw = self._mc.create_window((0, 0), window=self._mf, anchor="nw")
        self._mf.bind("<Configure>", lambda e: self._mc.config(scrollregion=self._mc.bbox("all")))
        self._mc.bind("<Configure>", lambda e: self._mc.itemconfig(self._mw, width=e.width))
        self._mc.bind("<MouseWheel>", lambda e:
            self._mc.yview_scroll(int(-1 * e.delta / 120), "units"))

        # ── Panels ────────────────────────────────────────────────────────
        info_card = tk.Frame(self._mf, bg=PANEL_BG)
        info_card.pack(fill="x", padx=16, pady=(16, 8))
        self.info_panel = InfoPanel(info_card)
        self.info_panel.pack(fill="x")
        self.info_panel.update_groups_tags(self._tags_groups)

        tk.Frame(self._mf, bg=BORDER_CLR, height=1).pack(fill="x", padx=16)

        # Options + Validation side by side
        mid_row = tk.Frame(self._mf, bg=BG)
        mid_row.pack(fill="x", padx=16, pady=(8, 8))
        mid_row.columnconfigure(0, weight=4)
        mid_row.columnconfigure(1, weight=2)   # validation panel width

        opt_card = tk.Frame(mid_row, bg=PANEL_BG)
        opt_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self.opts_panel = OptionsPanel(opt_card, on_change=self._auto_apply)
        self.opts_panel.pack(fill="x", padx=4)

        val_card = tk.Frame(mid_row, bg=PANEL_BG, bd=1, relief="solid")
        val_card.grid(row=0, column=1, sticky="nsew")
        self.val_panel = ValidationPanel(val_card)
        self.val_panel.pack(fill="both", expand=True)

        tk.Frame(self._mf, bg=BORDER_CLR, height=1).pack(fill="x", padx=16)

        tbl_card = tk.Frame(self._mf, bg=BG)
        tbl_card.pack(fill="both", expand=True, padx=16, pady=(8, 16))
        self.pivot = TabTable(tbl_card,
                                ingr_master=flat_ingredients(self._ingr_data),
                                on_data_change=None,
                                on_auto_sugar=self._auto_check_std_sugars)
        self.pivot.pack(fill="both", expand=True)

    # ─── Auto-apply ──────────────────────────────────────────────────────

    def _auto_apply(self):
        if self._current is None:
            return
        rd      = self._recipes[self._current]
        combos  = self.opts_panel.get_active_combos()
        sugars          = self.opts_panel.get_active_sugars()
        options_by_size = {sz: self.opts_panel.get_active_options_for_size(sz) for sz in SIZES}

        # Merge current rows into full cache so unchecked rows are preserved
        for r in rd.rows:
            k = (r.get("Size"), r.get("Option"), r.get("Sugar Level"))
            self._row_cache.setdefault(rd.name, {})[k] = r

        # Rebuild row list from cache (keeps data even after uncheck→recheck)
        cache = self._row_cache.get(rd.name, {})
        new_rows = []
        for (sz, op, sg) in combos:
            k = (sz, op, sg)
            if k in cache:
                new_rows.append(cache[k])
            else:
                row = {"Size": sz, "Option": op, "Sugar Level": sg, POPUP_HDR: ""}
                cache[k] = row
                new_rows.append(row)
        rd.rows = new_rows
        self.pivot.load(rd, options_by_size, sugars)
        self.val_panel.update_master(flat_ingredients(self._ingr_data))
        self.val_panel.run(self._recipes)
        self._take_snapshot()

    # ─── Recipe management ───────────────────────────────────────────────

    def _new_recipe_set(self):
        self._commit_current()
        rd = RecipeData("New Recipe")
        self._recipes   = [rd]
        self._file_path = None
        self._file_lbl.config(text="Unsaved")
        self._current   = 0
        self._undo_mgr  = UndoManager()
        self._row_cache = {}
        self._refresh_sidebar()
        self._load_recipe(0)
        self._take_snapshot()
        self._dirty = False   # fresh state = not dirty

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Open Excel Recipe File",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            self._commit_current()
            self._recipes   = load_workbook_data(path)
            self._file_path = path
            self._file_lbl.config(text=_short_name(os.path.basename(path)))
            self._current   = 0
            self._undo_mgr  = UndoManager()
            self._row_cache = {}
            # Pre-populate cache from loaded data
            for rd in self._recipes:
                for r in rd.rows:
                    k = (r.get("Size"), r.get("Option"), r.get("Sugar Level"))
                    self._row_cache.setdefault(rd.name, {})[k] = r
            self._refresh_sidebar()
            self._load_recipe(0)
            self._take_snapshot()
            self._dirty = False   # just opened = not dirty
        except Exception as ex:
            messagebox.showerror("Error", f"Could not open file:\n{ex}")

    def _import_file(self):
        """Import recipes from another Excel file and merge into the current session."""
        path = filedialog.askopenfilename(
            title="Import Recipes from File",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            self._commit_current()
            imported = load_workbook_data(path)
            if not imported:
                messagebox.showinfo("Import", "No recipes found in the selected file.")
                return
            # Append imported recipes to current list
            existing_names = {rd.name for rd in self._recipes}
            added = 0
            for rd in imported:
                # If name already exists, append a suffix to avoid collision
                name = rd.name
                if name in existing_names:
                    base = name
                    counter = 2
                    while name in existing_names:
                        name = f"{base} ({counter})"
                        counter += 1
                    rd.name = name
                existing_names.add(rd.name)
                self._recipes.append(rd)
                # Pre-populate row cache
                for r in rd.rows:
                    k = (r.get("Size"), r.get("Option"), r.get("Sugar Level"))
                    self._row_cache.setdefault(rd.name, {})[k] = r
                added += 1
            # Select the first newly imported recipe
            self._current = len(self._recipes) - added
            self._refresh_sidebar()
            self._load_recipe(self._current)
            self._take_snapshot()
            messagebox.showinfo("Import", f"Successfully imported {added} recipe(s) from:\n{os.path.basename(path)}")
        except Exception as ex:
            messagebox.showerror("Error", f"Could not import file:\n{ex}")

    def _save_file(self):
        if not self._file_path:
            self._save_as_file(); return
        self._do_save(self._file_path)

    def _save_as_file(self):
        path = filedialog.asksaveasfilename(
            title="Save As", defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self._file_path = path
        self._file_lbl.config(text=_short_name(os.path.basename(path)))
        self._do_save(path)

    def _do_save(self, path):
        self._commit_current()
        if self.pivot.has_validation_errors():
            messagebox.showerror("Validation Error",
                                 "Some Usage fields contain non-numeric values.\n"
                                 "Please fix the fields highlighted in red before saving.")
            return
        try:
            save_workbook(self._recipes, path)
            self._dirty = False
            messagebox.showinfo("Saved", f"File saved:\n{path}")
        except Exception as ex:
            messagebox.showerror("Save Error", str(ex))

    def _add_recipe(self):
        self._commit_current()
        rd = RecipeData(f"Recipe {len(self._recipes) + 1}")
        self._recipes.append(rd)
        self._current = len(self._recipes) - 1
        self._refresh_sidebar()
        self._load_recipe(self._current)
        self._take_snapshot()

    def _delete_recipe(self, idx):
        # Accept either a single index or a list of indices (multi-selection)
        if isinstance(idx, (list, tuple, set)):
            idxs = sorted(set(int(i) for i in idx), reverse=True)
        else:
            idxs = [int(idx)]
        if not idxs:
            return
        if len(self._recipes) - len(idxs) < 1:
            messagebox.showinfo("Info", "At least one recipe must remain.")
            return
        if len(idxs) == 1:
            msg = f"Delete recipe '{self._recipes[idxs[0]].name}'?"
        else:
            names = [self._recipes[i].name for i in sorted(idxs)]
            preview = "\n".join(f"  • {n}" for n in names[:6])
            more = f"\n  ...and {len(names) - 6} more" if len(names) > 6 else ""
            msg = f"Delete {len(idxs)} recipes?\n\n{preview}{more}"
        if not messagebox.askyesno("Delete", msg):
            return
        for i in idxs:  # already descending
            self._recipes.pop(i)
        self._current = max(0, min(idxs) - 1) if self._recipes else 0
        self.sidebar._multi_sel = set()
        self.sidebar._sel_anchor = None
        self._refresh_sidebar()
        self._load_recipe(self._current)
        self._take_snapshot()

    def _duplicate_recipe(self, idx):
        # Accept either a single index or a list of indices
        if isinstance(idx, (list, tuple, set)):
            idxs = sorted(set(int(i) for i in idx))
        else:
            idxs = [int(idx)]
        if not idxs:
            return
        self._commit_current()
        # Insert duplicates right after the last selected index so they appear together
        insert_at = idxs[-1] + 1
        new_indices = []
        for offset, src in enumerate(idxs):
            orig    = self._recipes[src]
            copy_rd = orig.deep_copy()
            copy_rd.name = orig.name + " (copy)"
            self._recipes.insert(insert_at + offset, copy_rd)
            new_indices.append(insert_at + offset)
        # Select the newly created duplicates as a group
        self._current = new_indices[0]
        self.sidebar._multi_sel = set(new_indices)
        self.sidebar._sel_anchor = new_indices[0]
        self._refresh_sidebar()
        self._load_recipe(self._current)
        self._take_snapshot()

    def _select_recipe(self, idx):
        self._commit_current()
        self._current = idx
        # Sidebar already updated _multi_sel via its click handler; just re-render.
        self._refresh_sidebar()
        self._load_recipe(idx)

    def _commit_current(self):
        if self._current is None or self._current >= len(self._recipes):
            return
        rd = self._recipes[self._current]
        self.info_panel.save(rd)
        self.pivot.collect(rd)

    def _load_recipe(self, idx):
        rd = self._recipes[idx]
        self._title_lbl.config(text=rd.name or "Untitled")
        self.info_panel.load(rd)
        self.opts_panel.load(rd)
        sugars          = self.opts_panel.get_active_sugars()
        options_by_size = {sz: self.opts_panel.get_active_options_for_size(sz) for sz in SIZES}
        # Fallback: infer from row data when nothing is checked (e.g. after loading file)
        if not any(options_by_size.values()):
            all_opts = [o for o in OPTIONS if any(r.get("Option") == o for r in rd.rows)]
            options_by_size = {sz: [o for o in all_opts
                                    if any(r.get("Size")==sz and r.get("Option")==o
                                           for r in rd.rows)]
                               for sz in SIZES}
        if not sugars:
            sugars = sorted(set(str(r.get("Sugar Level","")) for r in rd.rows) - {""}, key=sugar_key)
        self.pivot.load(rd, options_by_size, sugars)
        self.val_panel.update_master(flat_ingredients(self._ingr_data))
        self.val_panel.run(self._recipes)
        self._update_undo_btns()

    # ─── Undo / Redo ─────────────────────────────────────────────────────

    def _take_snapshot(self):
        self._commit_current()
        self._undo_mgr.snapshot(self._recipes)
        self._dirty = True
        self._update_undo_btns()

    def _undo(self):
        self._commit_current()
        snap = self._undo_mgr.undo()
        if snap:
            self._recipes = snap
            self._current = min(self._current or 0, len(self._recipes) - 1)
            self._refresh_sidebar()
            self._load_recipe(self._current)
        self._update_undo_btns()

    def _redo(self):
        self._commit_current()
        snap = self._undo_mgr.redo()
        if snap:
            self._recipes = snap
            self._current = min(self._current or 0, len(self._recipes) - 1)
            self._refresh_sidebar()
            self._load_recipe(self._current)
        self._update_undo_btns()

    def _update_undo_btns(self):
        self._undo_btn.config(state="normal" if self._undo_mgr.can_undo() else "disabled")
        self._redo_btn.config(state="normal" if self._undo_mgr.can_redo() else "disabled")

    # ─── Ingredient manager ───────────────────────────────────────────────

    def _open_ingr_mgr(self):
        def on_save(new_data):
            self._ingr_data = new_data
            self.pivot.update_ingr_master(flat_ingredients(new_data))
            self.val_panel.update_master(flat_ingredients(new_data))
            self.val_panel.run(self._recipes)
        IngredientManagerDialog(self, self._ingr_data, on_save)

    def _open_code_mgr(self):
        def on_save():
            # Re-run validation after mappings change (unknown codes may now resolve)
            self.val_panel.run(self._recipes)
        CodeMappingDialog(self, self._recipes, on_save)

    def _auto_check_std_sugars(self):
        """Fix 6: auto-check all standard sugar checkboxes."""
        for sg in STD_SUGARS:
            if sg in self.opts_panel._sugar_vars:
                self.opts_panel._sugar_vars[sg].set(True)
        self._auto_apply()

    def _open_groups_tags_mgr(self):
        """Open Groups & Tags manager dialog."""
        def on_save(tags_groups):
            self._tags_groups = tags_groups
            save_groups_tags(tags_groups)
            self.info_panel.update_groups_tags(tags_groups)
        GroupsTagsDialog(self, self._tags_groups, on_save)

    def _on_close(self):
        """Handle window close — warn if unsaved changes."""
        if self._dirty:
            ans = messagebox.askyesnocancel(
                "Unsaved Changes",
                "You have unsaved changes.\nDo you want to save before closing?"
            )
            if ans is None:          # Cancel → stay open
                return
            if ans:                  # Yes → save then close
                self._save_file()
                if self._dirty:      # Save was cancelled or failed → stay open
                    return
        self.destroy()

    def _export_qr(self):
        """Export QR Code Excel file from current loaded recipes."""
        self._commit_current()

        if not self._recipes or all(not rd.rows for rd in self._recipes):
            messagebox.showwarning("Export QR", "No recipe data to export.")
            return

        # Build default filename: "QR Code_<original filename>"
        if self._file_path:
            orig_name   = os.path.splitext(os.path.basename(self._file_path))[0]
            default_name = f"QR Code_{orig_name}.xlsx"
            init_dir     = os.path.dirname(self._file_path)
        else:
            default_name = "QR Code_export.xlsx"
            init_dir     = os.path.expanduser("~")

        out_path = filedialog.asksaveasfilename(
            title="Export QR Code File",
            initialfile=default_name,
            initialdir=init_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not out_path:
            return

        try:
            warnings = export_qr_workbook(
                self._recipes, out_path,
                flat_ingredients(self._ingr_data)
            )
            if warnings:
                warn_text = "\n".join(warnings[:20])
                if len(warnings) > 20:
                    warn_text += f"\n... and {len(warnings)-20} more"
                messagebox.showwarning(
                    "Export QR — Warnings",
                    f"File saved, but {len(warnings)} mapping issue(s) found:\n\n{warn_text}"
                )
            else:
                messagebox.showinfo("Export QR", f"QR file saved:\n{out_path}")
        except Exception as ex:
            messagebox.showerror("Export QR Error", str(ex))

    def _reorder_recipes(self, from_idx, to_idx):
        """Move recipe at from_idx to to_idx. from_idx may be a list of indices."""
        # Group move: pull all selected items out, then re-insert at the drop target
        if isinstance(from_idx, (list, tuple, set)):
            src_idxs = sorted(set(int(i) for i in from_idx))
            if not src_idxs:
                return
            # Pull out (descending so indices stay valid)
            moved = []
            for i in sorted(src_idxs, reverse=True):
                moved.append(self._recipes.pop(i))
            moved.reverse()  # restore original order
            # Adjust destination: how many removed items were before to_idx?
            shift = sum(1 for i in src_idxs if i < to_idx)
            insert_at = max(0, min(len(self._recipes), to_idx - shift))
            # If the user dropped on one of the items they were dragging,
            # land at the first item's original position (after compensation)
            if to_idx in src_idxs:
                insert_at = max(0, min(len(self._recipes), src_idxs[0] - sum(1 for i in src_idxs if i < src_idxs[0])))
            for offset, rd in enumerate(moved):
                self._recipes.insert(insert_at + offset, rd)
            new_indices = list(range(insert_at, insert_at + len(moved)))
            # Keep primary selection on the recipe that was previously current,
            # if it was part of the group; otherwise pick the first moved item.
            if self._current in src_idxs:
                rel = src_idxs.index(self._current)
                self._current = new_indices[rel]
            else:
                self._current = new_indices[0]
            self.sidebar._multi_sel = set(new_indices)
            self.sidebar._sel_anchor = new_indices[0]
            self._refresh_sidebar()
            self._take_snapshot()
            return
        # Single-item move (original behavior)
        if from_idx == to_idx:
            return
        rd = self._recipes.pop(from_idx)
        self._recipes.insert(to_idx, rd)
        # Adjust current selection
        if self._current == from_idx:
            self._current = to_idx
        elif from_idx < self._current <= to_idx:
            self._current -= 1
        elif to_idx <= self._current < from_idx:
            self._current += 1
        # Keep multi-selection consistent: just the moved item
        self.sidebar._multi_sel = {to_idx}
        self.sidebar._sel_anchor = to_idx
        self._refresh_sidebar()
        self._take_snapshot()

    def _refresh_sidebar(self):
        names = [rd.name or f"Recipe {i+1}" for i, rd in enumerate(self._recipes)]
        # Preserve any multi-selection set by the App/Sidebar
        multi = self.sidebar._multi_sel if self.sidebar._multi_sel else None
        self.sidebar.populate(names, self._current, multi)

# ─── Entry point ──────────────────────────────────────────────────────────────

def _short_name(name, max_len=32):
    """Truncate a filename for display in toolbar."""
    return name if len(name) <= max_len else name[:max_len - 3] + "..."

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        import sys
        print("Please install openpyxl:  pip install openpyxl")
        sys.exit(1)
    App().mainloop()
